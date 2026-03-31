from openpyxl import load_workbook
import re
path = r"c:\Users\dekok\Downloads\Q3 REPORTS (1)\Q3 REPORTS\Q3 Makgabaneng\Quarter 3 2025 Makgabaneng NCD  REPORT .xlsx"
wb=load_workbook(path, data_only=True, read_only=True)
rows=list(wb['TOTAL'].iter_rows(values_only=True))
for idx,row in enumerate(rows, start=1):
    code=str(row[1] if len(row)>1 and row[1] is not None else '').strip()
    name=str(row[2] if len(row)>2 and row[2] is not None else '').strip()
    if re.fullmatch(r'\d+[a-zA-Z]?|\d+', code) and name:
        print(idx, code, '|', name)
