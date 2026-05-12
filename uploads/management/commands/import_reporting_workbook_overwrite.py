import re
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import connection, transaction

from aggregates.models import Aggregate
from indicators.models import Indicator
from openpyxl import load_workbook
from organizations.models import Organization
from projects.models import Project, ProjectIndicator


SKIP_SHEET_ALIASES = {
    "indicator matrix",
    "total",
    "totals",
    "drop down",
    "dropdown",
    "instructions",
}
IGNORED_ORG_TOKENS = {"the", "of", "for", "and", "org", "organization"}
DEFAULT_AGE_BAND_BY_COLUMN = {
    "G": "10-14",
    "H": "10-14",
    "I": "10-14",
    "J": "10-14",
    "K": "10-14",
    "L": "15-19",
    "M": "15-19",
    "N": "15-19",
    "O": "15-19",
    "P": "15-19",
    "Q": "20-24",
    "R": "25-29",
    "S": "30-34",
    "T": "35-39",
    "U": "40-44",
    "V": "45-49",
    "W": "50-54",
    "X": "55-59",
    "Y": "60-64",
    "Z": "65+",
}
BONELA_AGE_BAND_BY_COLUMN = {
    "G": "10",
    "H": "11",
    "I": "12",
    "J": "13",
    "K": "14",
    "L": "15",
    "M": "16",
    "N": "17",
    "O": "18",
    "P": "19",
    "Q": "20-24",
    "R": "25-29",
    "S": "30-34",
    "T": "35-39",
    "U": "40-44",
    "V": "45-49",
    "W": "50-54",
    "X": "55-59",
    "Y": "60-64",
    "Z": "65+",
}
AGE_COLUMNS = tuple(DEFAULT_AGE_BAND_BY_COLUMN.keys())
TOTAL_COLUMNS = ("AA", "F", "S", "T", "G", "R", "Q", "P", "O", "N", "M", "L", "K", "J", "I", "H")
ORG_ALIASES = {
    "masego mental health org": "masego mental health",
    "the just hope foundation": "just hope foundation",
    "vmf valour": "valour mental health",
    "stop smoking support group": "sssg",
    "the fighters support group": "tfsg",
    "botswana network for mental health bonmeh": "bonmeh",
    "botswana association for the bl": "botswana association for the blind and partially sighted",
    "botswana association for the blind": "botswana association for the blind and partially sighted",
    "botswana association of the dea": "botswana association for the deaf",
    "botswana council of the disable": "bcd",
    "botswana council of the disabled": "bcd",
    "botswana society for the disabl": "botswana society for the disabled",
    "kebotlhokwa": "kebotlhekwa",
    "lesbians gays bisexuals of b": "legabibo",
}
SECTION_INDEX_PATTERN = re.compile(r"^\d+[a-z]?$", re.IGNORECASE)
COUNT_PREFIXES = (
    "total number of ",
    "number of ",
    "total ",
)
PRIMARY_SUB_LABEL_VALUE_SETS = {
    "Key Population": {
        "general pop.",
        "general pop",
        "general population",
        "fsw",
        "msm",
        "pwid",
        "pwids",
        "pwd",
        "lgbtqi+",
        "lgbtq+",
    },
    "Alcohol Use": {"alcohol use"},
    "Tobacco Use": {"tobacco use"},
    "Family Planning": {
        "condom counselling",
        "oral contraceptives",
        "injectables",
        "implants",
        "iud",
        "referral",
    },
    "Community Leaders": {
        "traditional leaders",
        "religious leaders",
        "community leaders",
    },
    "Social Media Platform": {
        "facebook",
        "whatsapp",
        "twitter",
        "instagram",
        "tiktok",
        "radio",
        "tv",
        "community media",
        "physical",
        "printed media",
    },
    "Non Traditional Sites": {
        "bars",
        "taxi ranks",
        "schools",
        "markets",
        "churches",
        "community events",
        "sports events",
    },
    "NCD Prevention Messages": {
        "alcohol reduction messages",
        "tobacco control messages",
        "physical activity messages",
        "healthy diet messages",
        "weight management messages",
        "blood pressure messages",
        "blood glucose messages",
        "waist circumference messages",
        "psychosocial",
    },
    "Mental Health Management/Treatment Services": {
        "clinical care",
        "psychosocial support",
        "rehabilitation",
        "crisis services",
        "specialized care",
    },
    "Counselling Sessions": {
        "individual counselling",
        "group counselling",
        "family counselling",
        "peer counselling",
        "follow-up session",
    },
    "Mental Health Screening": {
        "suicide",
        "depression",
        "anxiety",
        "substance use disorders",
        "psychosocial",
        "living with a person with a mental health illness",
        "sleep hygiene",
        "others",
    },
    "NCD Screening": {
        "blood glucose",
        "blood pressure",
        "bmi",
        "waist circumference",
    },
    "Condom Type": {
        "male condom",
        "female condom",
        "braille-labelled condom",
        "brailled condom",
        "braille labeled condom",
    },
    "Activity Type": {
        "advocacy",
        "community dialogue",
        "campaign",
        "engagement",
        "training",
        "workshop",
        "meeting",
    },
    "Referral Type": {
        "hiv testing",
        "sti treatment",
        "gbv services",
        "ncd screening",
        "mental health",
        "family planning",
    },
    "Location": {"urban", "rural", "peri urban", "community", "facility"},
    "Target Group": {
        "youth",
        "adults",
        "adolescents",
        "general population",
        "key population",
        "people living with hiv",
    },
}
PREFERRED_SUB_LABEL_ORDER = [
    "Key Population",
    "Alcohol Use",
    "Tobacco Use",
    "Family Planning",
    "Community Leaders",
    "Social Media Platform",
    "Non Traditional Sites",
    "NCD Prevention Messages",
    "Mental Health Management/Treatment Services",
    "Counselling Sessions",
    "Mental Health Screening",
    "NCD Screening",
    "Condom Type",
    "Activity Type",
    "Referral Type",
    "Location",
    "Target Group",
    "Sex",
    "Age Range",
    "Disaggregate",
]
INDICATOR_TOKEN_STOPWORDS = {
    "a",
    "an",
    "and",
    "by",
    "for",
    "from",
    "in",
    "of",
    "on",
    "per",
    "the",
    "to",
    "total",
    "number",
    "people",
    "person",
    "population",
    "populations",
    "messages",
    "message",
    "reached",
    "with",
    "who",
    "were",
    "was",
    "is",
}
SEX_VALUE_ORDER = ["Male", "Female", "Other", "Unknown", "All"]
KEY_POPULATION_VALUE_ORDER = [
    "GENERAL POP.",
    "FSW",
    "MSM",
    "PWID",
    "PWUD",
    "PWD",
    "LGBTQI+",
    "All",
]
BONELA_SCOPE_TOKENS = {"bonela"}


def normalize_text(value) -> str:
    normalized = re.sub(r"[^a-z0-9]+", " ", str(value or "").lower()).strip()
    normalized = (
        normalized.replace("peopel", "people")
        .replace("elligible", "eligible")
        .replace("reffered", "referred")
        .replace("pyschosocial", "psychosocial")
        .replace("descrimination", "discrimination")
        .replace("coodinators", "coordinators")
        .replace("perforamance", "performance")
        .replace("identifies needs", "identified needs")
        .replace("field visists", "field visits")
        .replace("virsual presentations", "visual presentations")
        .replace("redness", "redress")
    )
    return re.sub(r"\s+", " ", normalized)


def canonical_indicator_name(value) -> str:
    normalized = normalize_text(value)
    previous = None
    while normalized and previous != normalized:
        previous = normalized
        for prefix in COUNT_PREFIXES:
            if normalized.startswith(prefix):
                normalized = normalized[len(prefix):].strip()
    normalized = normalized.replace("kvps", "key populations")
    normalized = normalized.replace("kvp", "key population")
    normalized = normalized.replace("people living with hiv", "plwh")
    normalized = normalized.replace("persons living with hiv", "plwh")
    normalized = normalized.replace("number of number of people", "number of people")
    normalized = normalized.replace("total number of", "number of")
    normalized = normalized.replace("number of number of", "number of")
    return re.sub(r"\s+", " ", normalized).strip()


def normalize_section_index(value) -> str:
    return re.sub(r"[^0-9a-z]+", "", str(value or "").lower()).strip()


def extract_section_index_from_code(value: str | None) -> str | None:
    raw_value = str(value or "").strip().lower()
    if not raw_value:
        return None

    normalized = normalize_section_index(raw_value)
    if SECTION_INDEX_PATTERN.fullmatch(normalized):
        return normalized

    tokens = re.findall(r"\d+[a-z]?", raw_value)
    if not tokens:
        return None

    short_tokens = [token for token in tokens if len(token) <= 3]
    return short_tokens[-1] if short_tokens else tokens[-1]


def is_auto_indicator(indicator: Indicator) -> bool:
    return str(indicator.code or "").upper().startswith("AUTO_")


def infer_primary_sub_label_from_value(value: dict | None) -> str | None:
    if not isinstance(value, dict):
        return None

    raw_disaggregates = value.get("disaggregates")
    normalized_values = set()
    if isinstance(raw_disaggregates, dict):
        normalized_values.update(
            normalize_text(primary_label)
            for primary_label in raw_disaggregates.keys()
            if primary_label not in (None, "", "All")
        )

    raw_categories = value.get("categories")
    if isinstance(raw_categories, dict):
        normalized_values.update(
            normalize_text(category_label)
            for category_label in raw_categories.keys()
            if category_label not in (None, "", "All")
        )

    if not normalized_values:
        return None

    best_label = None
    best_overlap = 0
    for label, known_values in PRIMARY_SUB_LABEL_VALUE_SETS.items():
        overlap = len(normalized_values & known_values)
        if overlap > best_overlap:
            best_overlap = overlap
            best_label = label

    if best_overlap:
        return best_label

    return None


def build_ordered_sub_labels(detected_disaggregations: set[str], value: dict | None = None) -> list[str]:
    labels = []
    inferred_primary = infer_primary_sub_label_from_value(value)
    if inferred_primary:
        labels.append(inferred_primary)

    normalized_detected = {normalize_text(item) for item in detected_disaggregations if item}
    explicit_map = {
        "kp": "Key Population",
        "key population": "Key Population",
        "alcohol use": "Alcohol Use",
        "tobacco use": "Tobacco Use",
        "family planning": "Family Planning",
        "community leaders": "Community Leaders",
        "social media platform": "Social Media Platform",
        "ncd prevention messages": "NCD Prevention Messages",
        "mental health management treatment services": "Mental Health Management/Treatment Services",
        "mental health management/treatment services": "Mental Health Management/Treatment Services",
        "counselling sessions": "Counselling Sessions",
        "counseling sessions": "Counselling Sessions",
        "receiving counselling": "Mental Health Screening",
        "receiving counseling": "Mental Health Screening",
        "mental health screening": "Mental Health Screening",
        "ncd screening": "NCD Screening",
        "condom type": "Condom Type",
        "activity type": "Activity Type",
        "referral type": "Referral Type",
        "location": "Location",
        "target group": "Target Group",
        "sex": "Sex",
        "age group": "Age Range",
        "age range": "Age Range",
        "service category": "Disaggregate",
    }

    for raw_label in normalized_detected:
        mapped_label = explicit_map.get(raw_label)
        if mapped_label == "Disaggregate" and inferred_primary:
            continue
        if mapped_label and mapped_label not in labels:
            labels.append(mapped_label)

    order_index = {label: index for index, label in enumerate(PREFERRED_SUB_LABEL_ORDER)}
    return sorted(labels, key=lambda label: (order_index.get(label, len(order_index)), label))


def significant_indicator_tokens(value: str) -> set[str]:
    return {
        token
        for token in canonical_indicator_name(value).split()
        if token and token not in INDICATOR_TOKEN_STOPWORDS
    }


def normalize_dimension_key(value: str) -> str:
    return normalize_text(value).replace(" ", "_")


def age_band_sort_key(value: str) -> tuple[int, int, str]:
    text = str(value or "").strip()
    if not text:
        return (10_000, 10_000, text)

    single_match = re.fullmatch(r"(\d{1,2})", text)
    if single_match:
        start = int(single_match.group(1))
        return (start, start, text)

    range_match = re.fullmatch(r"(\d{1,2})\s*-\s*(\d{1,2})", text)
    if range_match:
        start = int(range_match.group(1))
        end = int(range_match.group(2))
        return (start, end, text)

    plus_match = re.fullmatch(r"(\d{1,2})\s*\+", text)
    if plus_match:
        start = int(plus_match.group(1))
        return (start, 999, text)

    return (10_000, 10_000, text.lower())


def ordered_unique(values: list[str]) -> list[str]:
    seen = set()
    ordered = []
    for value in values:
        normalized = str(value or "").strip()
        if not normalized or normalized in seen:
            continue
        seen.add(normalized)
        ordered.append(normalized)
    return ordered


def normalize_age_band_label(value: str | None) -> str | None:
    text = str(value or "").strip()
    if not text:
        return None

    normalized = re.sub(r"\s+", "", text.lower())
    if re.fullmatch(r"\d{1,2}", normalized):
        return str(int(normalized))
    if re.fullmatch(r"\d{1,2}-\d{1,2}", normalized):
        start, end = normalized.split("-", 1)
        return f"{int(start)}-{int(end)}"
    if re.fullmatch(r"\d{1,2}\+", normalized):
        return f"{int(normalized[:-1])}+"
    return None


def is_non_age_summary_label(value: str | None) -> bool:
    normalized = normalize_text(value)
    return normalized in {"sub total", "subtotal", "total", "ayp 10 24", "ayp", "value"}


def sort_dimension_values(values: list[str], label: str) -> list[str]:
    ordered = ordered_unique(values)
    normalized_label = normalize_text(label)
    if normalized_label in {"age range", "age band"}:
        return sorted(ordered, key=age_band_sort_key)
    if normalized_label == "sex":
        order_map = {value.lower(): index for index, value in enumerate(SEX_VALUE_ORDER)}
        return sorted(ordered, key=lambda value: (order_map.get(value.lower(), 999), value.lower()))
    if normalized_label == "key population":
        order_map = {value.lower(): index for index, value in enumerate(KEY_POPULATION_VALUE_ORDER)}
        return sorted(ordered, key=lambda value: (order_map.get(value.lower(), 999), value.lower()))
    return ordered


def build_disaggregation_config(value: dict | None, desired_sub_labels: list[str]) -> dict:
    if not isinstance(value, dict):
        return {}

    raw_disaggregates = value.get("disaggregates")
    if not isinstance(raw_disaggregates, dict) or not raw_disaggregates:
        return {}

    primary_values = ordered_unique(
        [str(primary).strip() for primary in raw_disaggregates.keys() if str(primary or "").strip()]
    )

    secondary_values = []
    band_values = []
    for secondary_map in raw_disaggregates.values():
        if not isinstance(secondary_map, dict):
            continue
        secondary_values.extend(
            [str(secondary).strip() for secondary in secondary_map.keys() if str(secondary or "").strip()]
        )
        for band_map in secondary_map.values():
            if not isinstance(band_map, dict):
                continue
            band_values.extend([str(band).strip() for band in band_map.keys() if str(band or "").strip()])

    primary_values = ordered_unique(primary_values)
    secondary_values = sort_dimension_values(secondary_values, "Sex")
    band_values = sort_dimension_values(band_values, "Age Range")

    primary_label = next(
        (label for label in desired_sub_labels if label not in {"Sex", "Age Range"}),
        "Disaggregate",
    )
    dimensions = []

    if primary_values and not (len(primary_values) == 1 and primary_values[0] == "All"):
        dimensions.append(
            {
                "key": normalize_dimension_key(primary_label),
                "label": primary_label,
                "values": sort_dimension_values(primary_values, primary_label),
            }
        )

    if secondary_values and not (len(secondary_values) == 1 and secondary_values[0] == "All"):
        dimensions.append(
            {
                "key": "sex",
                "label": "Sex",
                "values": secondary_values,
            }
        )

    if band_values:
        dimensions.append(
            {
                "key": "age_band",
                "label": "Age Range",
                "values": band_values,
            }
        )

    if not dimensions:
        return {}

    if len(dimensions) == 1:
        layout = "list"
    elif len(dimensions) == 2:
        layout = "matrix"
    else:
        layout = "nested-matrix"

    return {
        "enabled": True,
        "layout": layout,
        "dimensions": dimensions,
    }


def merge_disaggregation_configs(existing: dict | None, candidate: dict | None) -> dict:
    existing = dict(existing or {})
    candidate = dict(candidate or {})

    if not candidate:
        return existing
    if not existing:
        return candidate
    if not existing.get("enabled"):
        return candidate
    if not candidate.get("enabled"):
        return existing

    existing_dimensions = list(existing.get("dimensions") or [])
    candidate_dimensions = list(candidate.get("dimensions") or [])
    if len(existing_dimensions) != len(candidate_dimensions):
        return existing

    merged_dimensions = []
    for existing_dimension, candidate_dimension in zip(existing_dimensions, candidate_dimensions):
        existing_key = normalize_text(existing_dimension.get("key"))
        existing_label = normalize_text(existing_dimension.get("label"))
        candidate_key = normalize_text(candidate_dimension.get("key"))
        candidate_label = normalize_text(candidate_dimension.get("label"))
        if existing_key != candidate_key or existing_label != candidate_label:
            return existing

        merged_values = ordered_unique(
            list(existing_dimension.get("values") or []) + list(candidate_dimension.get("values") or [])
        )
        merged_dimensions.append(
            {
                "key": existing_dimension.get("key"),
                "label": existing_dimension.get("label"),
                "values": sort_dimension_values(merged_values, existing_dimension.get("label") or ""),
            }
        )

    merged = dict(existing)
    merged["enabled"] = True
    merged["layout"] = existing.get("layout") or candidate.get("layout")
    merged["dimensions"] = merged_dimensions
    return merged


def sanitize_indicator_code(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9_-]+", "-", str(value or "").strip().upper()).strip("-_")
    return (cleaned or "INDICATOR")[:50]


def unique_indicator_code(seed: str) -> str:
    base = sanitize_indicator_code(seed)
    candidate = base
    suffix = 2
    while Indicator.objects.filter(code__iexact=candidate).exists():
        candidate = f"{base[: max(1, 50 - len(str(suffix)) - 1)]}-{suffix}"
        suffix += 1
    return candidate


def to_decimal(value) -> Decimal | None:
    if value in (None, ""):
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))

    text = str(value).strip()
    if not text:
        return None

    match = re.search(r"-?\d+(?:\.\d+)?", text.replace(",", ""))
    if not match:
        return None

    try:
        return Decimal(match.group(0))
    except (InvalidOperation, TypeError, ValueError):
        return None


def to_numeric_cell_decimal(value) -> Decimal | None:
    if value in (None, ""):
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))

    text = str(value).strip()
    if not text or text.startswith("="):
        return None
    if normalize_age_band_label(text) or is_non_age_summary_label(text):
        return None

    compact = text.replace(",", "")
    if not re.fullmatch(r"-?\d+(?:\.\d+)?", compact):
        return None

    try:
        return Decimal(compact)
    except (InvalidOperation, TypeError, ValueError):
        return None


def decimal_to_json_number(value: Decimal | None):
    if value is None:
        return None
    return int(value) if value == value.to_integral_value() else float(value)


def to_decimal_or_zero(value) -> Decimal:
    parsed = to_decimal(value)
    return parsed if parsed is not None else Decimal("0")


def merge_json_values(left, right):
    if isinstance(left, dict) and isinstance(right, dict):
        merged = dict(left)
        for key, value in right.items():
            if key in merged:
                merged[key] = merge_json_values(merged[key], value)
            else:
                merged[key] = value
        return merged
    if isinstance(left, (int, float, Decimal)) and isinstance(right, (int, float, Decimal)):
        return decimal_to_json_number(to_decimal_or_zero(left) + to_decimal_or_zero(right))
    return right if right not in (None, "") else left


def ensure_project_indicator(project: Project, indicator: Indicator):
    project_indicator = ProjectIndicator.objects.filter(project=project, indicator=indicator).first()
    if project_indicator:
        return project_indicator, False

    try:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                INSERT INTO projects_projectindicator
                    (project_id, indicator_id, target_value, current_value, baseline_value, q1_target, q2_target, q3_target, q4_target)
                VALUES (%s, %s, 0, 0, 0, 0, 0, 0, 0)
                ON CONFLICT (project_id, indicator_id) DO NOTHING
                """,
                [project.id, indicator.id],
            )
        project_indicator = ProjectIndicator.objects.get(project=project, indicator=indicator)
        return project_indicator, True
    except Exception:
        project_indicator, created = ProjectIndicator.objects.get_or_create(
            project=project,
            indicator=indicator,
            defaults={
                "target_value": 0,
                "current_value": 0,
                "baseline_value": 0,
            },
        )
        return project_indicator, created


def organization_matches_bonela(organization: Organization | None) -> bool:
    if not organization:
        return False

    candidate_values = [organization.name, organization.code]
    for value in candidate_values:
        normalized = normalize_text(value)
        if not normalized:
            continue
        if normalized == "bonela":
            return True
        if BONELA_SCOPE_TOKENS & set(normalized.split()):
            return True

    return False


def uses_bonela_single_year_age_bands(
    organization: Organization | None,
    coordinator: Organization | None = None,
) -> bool:
    if organization_matches_bonela(coordinator):
        return True

    current = organization
    seen_ids = set()
    while current and current.id not in seen_ids:
        if organization_matches_bonela(current):
            return True
        seen_ids.add(current.id)
        current = current.parent

    return False


def get_age_band_mapping(
    organization: Organization | None,
    coordinator: Organization | None = None,
) -> dict[str, str]:
    if uses_bonela_single_year_age_bands(organization, coordinator):
        return BONELA_AGE_BAND_BY_COLUMN
    return DEFAULT_AGE_BAND_BY_COLUMN


def sum_cells(row: dict, columns: tuple[str, ...]) -> Decimal:
    total = Decimal("0")
    for column in columns:
        value = to_decimal(row.get(column))
        if value is not None:
            total += value
    return total


def pick_total(row: dict) -> Decimal | None:
    for column in TOTAL_COLUMNS:
        candidate = to_numeric_cell_decimal(row.get(column))
        if candidate is not None:
            return candidate
    return None


def organization_variants(value: str) -> list[str]:
    normalized = normalize_text(value)
    variants = {normalized}
    if normalized in ORG_ALIASES:
        variants.add(normalize_text(ORG_ALIASES[normalized]))
    if normalized.startswith("the "):
        variants.add(normalized[4:])
    if normalized.endswith(" org"):
        variants.add(normalized[:-4].strip())
    return [variant for variant in variants if variant]


def normalized_sheet_name(value: str) -> str:
    return normalize_text(value)


def significant_organization_tokens(value: str) -> set[str]:
    return {
        token
        for token in normalize_text(value).split()
        if token and token not in IGNORED_ORG_TOKENS
    }


def is_skipped_sheet(sheet_name: str) -> bool:
    return normalized_sheet_name(sheet_name) in SKIP_SHEET_ALIASES


def find_matrix_sheet_name(sheet_names: list[str]) -> str | None:
    for sheet_name in sheet_names:
        if normalized_sheet_name(sheet_name) == "indicator matrix":
            return sheet_name
    return None


class OrganizationResolver:
    def __init__(self):
        self.organizations = list(Organization.objects.all())
        self.index = {}
        for organization in self.organizations:
            keys = {
                normalize_text(organization.name),
                normalize_text(organization.code),
            }
            keys.update(organization_variants(organization.name))
            self.index[organization.id] = {
                "organization": organization,
                "keys": keys,
                "tokens": significant_organization_tokens(organization.name),
            }

    def resolve(self, name: str) -> Organization | None:
        if not name:
            return None

        normalized_name = normalize_text(name)
        for variant in organization_variants(name):
            exact = [
                item["organization"]
                for item in self.index.values()
                if variant in item["keys"]
            ]
            if exact:
                return exact[0]

        prefix_matches = []
        for item in self.index.values():
            matched_prefix = any(
                key.startswith(normalized_name) or normalized_name.startswith(key)
                for key in item["keys"]
                if key
            )
            if matched_prefix:
                prefix_matches.append(item["organization"])

        if len(prefix_matches) == 1:
            return prefix_matches[0]

        requested_tokens = significant_organization_tokens(name)
        ranked = []
        for item in self.index.values():
            overlap = len(requested_tokens & item["tokens"])
            if overlap:
                ranked.append((overlap, len(item["tokens"]), item["organization"]))
        ranked.sort(key=lambda item: (-item[0], item[1], item[2].name.lower()))
        if not ranked:
            return None

        best_overlap, _, best_match = ranked[0]
        minimum_overlap = max(2, len(requested_tokens) // 2)
        return best_match if best_overlap >= minimum_overlap else None


class IndicatorResolver:
    def __init__(self, project: Project | None = None):
        self.project_indicator_ids = set()
        if project:
            self.project_indicator_ids = set(
                ProjectIndicator.objects.filter(project=project).values_list("indicator_id", flat=True)
            )

        self.indicators = []
        self.project_indicators = []
        self.configured_indicators = []
        self.configured_project_indicators = []
        self.indicators_by_index: dict[str, list[Indicator]] = {}
        self.project_indicators_by_index: dict[str, list[Indicator]] = {}
        for indicator in Indicator.objects.all():
            self.remember(indicator)

    def _candidate_sort_key(self, indicator: Indicator, requested_key: str, section_index: str | None):
        candidate_key = canonical_indicator_name(indicator.name)
        requested_tokens = set(requested_key.split())
        candidate_tokens = set(candidate_key.split())
        overlap = len(requested_tokens & candidate_tokens)
        requested_significant_tokens = significant_indicator_tokens(requested_key)
        candidate_significant_tokens = significant_indicator_tokens(candidate_key)
        significant_overlap = len(requested_significant_tokens & candidate_significant_tokens)
        is_exact = candidate_key == requested_key
        is_prefix_match = candidate_key.startswith(requested_key) or requested_key.startswith(candidate_key)
        index_matches = section_index and extract_section_index_from_code(indicator.code) == section_index
        return (
            0 if index_matches else 1,
            0 if is_exact else 1,
            0 if is_prefix_match else 1,
            -significant_overlap,
            -overlap,
            0 if indicator.id in self.project_indicator_ids else 1,
            0 if not is_auto_indicator(indicator) else 1,
            0 if indicator.sub_labels else 1,
            abs(len(candidate_key) - len(requested_key)),
            indicator.name.lower(),
        )

    def _best_exact(self, key: str, candidates: list[Indicator], section_index: str | None) -> Indicator | None:
        exact_matches = [
            indicator for indicator in candidates if canonical_indicator_name(indicator.name) == key
        ]
        if not exact_matches:
            return None
        return sorted(
            exact_matches,
            key=lambda indicator: self._candidate_sort_key(indicator, key, section_index),
        )[0]

    def _best_fuzzy(
        self,
        key: str,
        candidates: list[Indicator],
        section_index: str | None,
        minimum_overlap: int,
    ) -> Indicator | None:
        requested_tokens = set(key.split())
        ranked = []
        requested_significant_tokens = significant_indicator_tokens(key)
        for indicator in candidates:
            candidate_key = canonical_indicator_name(indicator.name)
            overlap = len(requested_tokens & set(candidate_key.split()))
            candidate_significant_tokens = significant_indicator_tokens(candidate_key)
            significant_overlap = len(requested_significant_tokens & candidate_significant_tokens)
            if overlap < minimum_overlap:
                continue
            if requested_significant_tokens and significant_overlap == 0:
                continue
            ranked.append(
                (
                    *self._candidate_sort_key(indicator, key, section_index),
                    indicator,
                )
            )

        if not ranked:
            return None

        ranked.sort(key=lambda item: item[:-1])
        return ranked[0][-1]

    def resolve(self, title: str, section_index: str | None = None) -> Indicator | None:
        key = canonical_indicator_name(title)
        normalized_index = normalize_section_index(section_index)

        exact_project_match = self._best_exact(key, self.configured_project_indicators, normalized_index)
        if exact_project_match:
            return exact_project_match

        exact_global_match = self._best_exact(key, self.configured_indicators, normalized_index)
        if exact_global_match:
            return exact_global_match

        exact_any_project_match = self._best_exact(key, self.project_indicators, normalized_index)
        if exact_any_project_match:
            return exact_any_project_match

        exact_any_global_match = self._best_exact(key, self.indicators, normalized_index)
        if exact_any_global_match:
            return exact_any_global_match

        if normalized_index:
            index_candidates = [
                indicator
                for indicator in self.project_indicators_by_index.get(normalized_index, [])
                if not is_auto_indicator(indicator)
            ]
            if index_candidates:
                best_match = self._best_fuzzy(
                    key,
                    index_candidates,
                    normalized_index,
                    minimum_overlap=max(2, len(set(key.split())) // 3),
                )
                if best_match:
                    return best_match

        fuzzy_project_match = self._best_fuzzy(
            key,
            self.configured_project_indicators,
            normalized_index,
            minimum_overlap=max(3, len(set(key.split())) // 2),
        )
        if fuzzy_project_match:
            return fuzzy_project_match

        fuzzy_global_match = self._best_fuzzy(
            key,
            self.configured_indicators,
            normalized_index,
            minimum_overlap=max(3, len(set(key.split())) // 2),
        )
        if fuzzy_global_match:
            return fuzzy_global_match

        fuzzy_any_project_match = self._best_fuzzy(
            key,
            self.project_indicators,
            normalized_index,
            minimum_overlap=max(3, len(set(key.split())) // 2),
        )
        if fuzzy_any_project_match:
            return fuzzy_any_project_match

        return self._best_fuzzy(
            key,
            self.indicators,
            normalized_index,
            minimum_overlap=max(3, len(set(key.split())) // 2),
        )

    def remember(self, indicator: Indicator):
        if indicator.id not in {item.id for item in self.indicators}:
            self.indicators.append(indicator)
            if not is_auto_indicator(indicator):
                self.configured_indicators.append(indicator)

        if indicator.id in self.project_indicator_ids and indicator.id not in {
            item.id for item in self.project_indicators
        }:
            self.project_indicators.append(indicator)
            if not is_auto_indicator(indicator):
                self.configured_project_indicators.append(indicator)

        section_index = extract_section_index_from_code(indicator.code)
        if section_index:
            self.indicators_by_index.setdefault(section_index, []).append(indicator)
            if indicator.id in self.project_indicator_ids:
                self.project_indicators_by_index.setdefault(section_index, []).append(indicator)


SECTION_SNAPSHOT_COLUMNS = ("B", "C", "E", "F", *AGE_COLUMNS, "AA")
SECTION_SNAPSHOT_INDEXES = {
    "B": 0,
    "C": 1,
    "E": 3,
    "F": 4,
    "G": 5,
    "H": 6,
    "I": 7,
    "J": 8,
    "K": 9,
    "L": 10,
    "M": 11,
    "N": 12,
    "O": 13,
    "P": 14,
    "Q": 15,
    "R": 16,
    "S": 17,
    "T": 18,
    "U": 19,
    "V": 20,
    "W": 21,
    "X": 22,
    "Y": 23,
    "Z": 24,
    "AA": 25,
}
MATRIX_ROW_INDEXES = {
    "A": 0,
    "I": 8,
    "J": 9,
    "K": 10,
    "L": 11,
    "N": 13,
}


def snapshot_row_values(row_values, formula_row_values=None) -> dict:
    values = tuple(row_values)
    formula_values = tuple(formula_row_values or ())
    snapshot = {
        column: values[index] if index < len(values) else None
        for column, index in SECTION_SNAPSHOT_INDEXES.items()
    }
    snapshot["__formula_source_columns"] = {
        column: bool(index < len(formula_values) and formula_values[index] not in (None, ""))
        for column, index in SECTION_SNAPSHOT_INDEXES.items()
        if column in AGE_COLUMNS
    }
    return snapshot


def row_contains_age_header(row: dict) -> bool:
    header_token = normalize_text(row.get("F"))
    if header_token not in {"age sex", "age"}:
        return False

    return (
        sum(
            1
            for column in AGE_COLUMNS
            if isinstance(row.get(column), str) and normalize_age_band_label(row.get(column))
        )
        >= 2
    )


def parse_matrix_sheet(ws) -> dict[str, dict]:
    assignments = {}
    current_org_name = None
    for row_values in ws.iter_rows(min_row=6, min_col=1, max_col=14, values_only=True):
        organization_name = str(row_values[MATRIX_ROW_INDEXES["A"]] or "").strip()
        if organization_name:
            current_org_name = organization_name

        indicator_title = str(row_values[MATRIX_ROW_INDEXES["N"]] or "").strip()
        if not current_org_name or not indicator_title:
            continue

        key = canonical_indicator_name(indicator_title)
        entry = assignments.setdefault(
            key,
            {
                "title": indicator_title,
                "assignments": [],
            },
        )
        entry["assignments"].append(
            {
                "organization_name": current_org_name,
                "q1_target": to_decimal(row_values[MATRIX_ROW_INDEXES["I"]]) or Decimal("0"),
                "q2_target": to_decimal(row_values[MATRIX_ROW_INDEXES["J"]]) or Decimal("0"),
                "q3_target": to_decimal(row_values[MATRIX_ROW_INDEXES["K"]]) or Decimal("0"),
                "q4_target": to_decimal(row_values[MATRIX_ROW_INDEXES["L"]]) or Decimal("0"),
            }
        )
    return assignments


def parse_sections(ws, formula_ws=None) -> list[dict]:
    sections = []
    current = None
    current_age_header = None
    formula_rows = (
        formula_ws.iter_rows(min_row=1, min_col=2, max_col=27, values_only=True)
        if formula_ws is not None
        else None
    )
    value_rows = ws.iter_rows(min_row=1, min_col=2, max_col=27, values_only=True)

    for row_values, formula_row_values in zip(value_rows, formula_rows or value_rows):
        row_snapshot = snapshot_row_values(row_values, formula_row_values)
        if row_contains_age_header(row_snapshot):
            current_age_header = row_snapshot
        index_value = str(row_snapshot["B"] or "").strip()
        title = str(row_snapshot["C"] or "").strip()
        is_section_start = bool(SECTION_INDEX_PATTERN.fullmatch(index_value))

        if is_section_start:
            if current:
                sections.append(current)
            current = {
                "index": index_value,
                "title": title,
                "header_row": current_age_header,
                "rows": [row_snapshot],
            }
            continue

        if current:
            if not current["title"] and title:
                current["title"] = title
            current["rows"].append(row_snapshot)

    if current:
        sections.append(current)
    return sections


def resolve_section_age_band_mapping(
    section: dict,
    fallback_mapping: dict[str, str],
) -> dict[str, str]:
    detected_mapping: dict[str, str] = {}

    candidate_rows = []
    header_row = section.get("header_row")
    if isinstance(header_row, dict):
        candidate_rows.append(header_row)
    candidate_rows.extend(section["rows"])

    for row in candidate_rows:
        header_token = normalize_text(row.get("F"))
        if header_token not in {"age sex", "age"}:
            continue

        row_mapping: dict[str, str] = {}
        saw_summary_label = False
        for column in AGE_COLUMNS:
            normalized_age_band = normalize_age_band_label(row.get(column))
            if normalized_age_band:
                row_mapping[column] = normalized_age_band
                continue
            if is_non_age_summary_label(row.get(column)):
                saw_summary_label = True
        if row_mapping:
            detected_mapping = row_mapping
            if saw_summary_label:
                break

    return detected_mapping or fallback_mapping


def extract_section_value(section: dict, age_band_by_column: dict[str, str]) -> tuple[dict, set[str]]:
    male_total = None
    female_total = None
    total_value = None
    categories = {}
    disaggregate_matrix = {}
    disaggregations = set()
    current_matrix_bucket = None
    summary_sex_block_active = False
    section_age_band_by_column = resolve_section_age_band_mapping(section, age_band_by_column)

    def normalize_summary_label(value: str) -> str:
        return normalize_text(value).replace("-", " ").strip()

    def is_summary_row(label: str) -> bool:
        normalized = normalize_summary_label(label)
        return normalized in {"total", "sub total", "subtotal", "total male", "total female"}

    def sum_category_payload(value) -> Decimal:
        if isinstance(value, dict):
            return sum((sum_category_payload(item) for item in value.values()), Decimal("0"))
        return to_decimal(value) or Decimal("0")

    def set_matrix_value(primary_label: str, sex_label: str, row: dict):
        nonlocal male_total, female_total
        if not primary_label:
            return

        sex_key = "Male" if sex_label in {"m", "male"} else "Female"
        band_payload = {}
        row_sum = Decimal("0")
        has_band_value = False

        for column, band_label in section_age_band_by_column.items():
            numeric_value = to_numeric_cell_decimal(row.get(column))
            if numeric_value is None:
                continue
            existing_band_total = to_decimal(band_payload.get(band_label)) or Decimal("0")
            number = existing_band_total + numeric_value
            band_payload[band_label] = decimal_to_json_number(number)
            row_sum += numeric_value
            has_band_value = True

        if not has_band_value:
            return

        bucket = disaggregate_matrix.setdefault(primary_label, {})
        bucket[sex_key] = band_payload
        disaggregations.update({"age group", "sex"})
        if primary_label != "All":
            disaggregations.add("service category")

        if sex_key == "Male":
            male_total = (male_total or Decimal("0")) + row_sum
        else:
            female_total = (female_total or Decimal("0")) + row_sum

    def set_age_only_matrix_value(primary_label: str, secondary_label: str, row: dict):
        if not primary_label:
            return

        dimension_key = secondary_label or "All"
        band_payload = {}
        has_band_value = False

        for column, band_label in section_age_band_by_column.items():
            numeric_value = to_numeric_cell_decimal(row.get(column))
            if numeric_value is None:
                continue
            existing_band_total = to_decimal(band_payload.get(band_label)) or Decimal("0")
            band_payload[band_label] = decimal_to_json_number(existing_band_total + numeric_value)
            has_band_value = True

        if not has_band_value:
            return

        bucket = disaggregate_matrix.setdefault(primary_label, {})
        existing_dimension = bucket.get(dimension_key)
        if existing_dimension and isinstance(existing_dimension, dict):
            for band_label, band_value in band_payload.items():
                existing_band_total = to_decimal(existing_dimension.get(band_label)) or Decimal("0")
                existing_dimension[band_label] = decimal_to_json_number(
                    existing_band_total + (to_decimal(band_value) or Decimal("0"))
                )
        else:
            bucket[dimension_key] = band_payload

        disaggregations.add("age group")
        if primary_label != "All":
            disaggregations.add("service category")

    for row in section["rows"]:
        sex = normalize_text(row.get("F"))
        primary_label = str(row.get("E") or "").strip()
        label = primary_label or str(row.get("F") or "").strip()
        is_summary_primary = bool(primary_label and is_summary_row(primary_label))
        is_summary_label = is_summary_row(label)
        section_total = pick_total(row)
        normalized_label = normalize_summary_label(label)
        has_age_values = any(
            to_numeric_cell_decimal(row.get(column)) is not None
            for column in section_age_band_by_column
        )
        formula_source_columns = row.get("__formula_source_columns") or {}
        has_age_structure = any(
            formula_source_columns.get(column)
            for column in section_age_band_by_column
        )

        if has_age_values:
            disaggregations.add("age group")

        if primary_label and not is_summary_primary:
            current_matrix_bucket = primary_label
            summary_sex_block_active = False
        elif primary_label and is_summary_primary:
            summary_sex_block_active = True

        if sex in {"m", "male", "f", "female"}:
            disaggregations.add("sex")
            if has_age_values and not is_summary_label and not summary_sex_block_active:
                matrix_bucket = current_matrix_bucket or "All"
                set_matrix_value(matrix_bucket, sex, row)
                continue

            if section_total is None:
                section_total = to_decimal(row.get("F"))
            if section_total is not None:
                if sex in {"m", "male"}:
                    male_total = section_total
                else:
                    female_total = section_total

        row_secondary_label = str(row.get("F") or "").strip()
        if (
            has_age_values
            and sex not in {"m", "male", "f", "female"}
            and not is_summary_row(label)
            and not is_non_age_summary_label(row.get("F"))
            and (primary_label or row_secondary_label)
        ):
            matrix_bucket = primary_label or current_matrix_bucket or "All"
            set_age_only_matrix_value(matrix_bucket, row_secondary_label or "All", row)
            continue

        if label and "total" in normalized_label:
            candidate_total = pick_total(row)
            if candidate_total is not None:
                total_value = candidate_total

        if label and not is_summary_row(label) and label != str(row.get("F") or "").strip():
            disaggregations.add("service category")
            if sex in {"m", "male", "f", "female"}:
                bucket = categories.setdefault(label, {})
                bucket["male" if sex in {"m", "male"} else "female"] = decimal_to_json_number(
                    section_total or Decimal("0")
                )
            elif section_total is not None:
                categories[label] = decimal_to_json_number(section_total)

    if disaggregate_matrix:
        recalculated_total = Decimal("0")
        recalculated_male = Decimal("0")
        recalculated_female = Decimal("0")
        for dimension_map in disaggregate_matrix.values():
            for sex_key, band_map in dimension_map.items():
                row_total = Decimal("0")
                for raw_value in band_map.values():
                    row_total += to_decimal(raw_value) or Decimal("0")
                recalculated_total += row_total
                if normalize_text(sex_key) == "male":
                    recalculated_male += row_total
                elif normalize_text(sex_key) == "female":
                    recalculated_female += row_total
        total_value = recalculated_total
        male_total = recalculated_male
        female_total = recalculated_female
    elif total_value is None:
        if male_total is not None or female_total is not None:
            total_value = (male_total or Decimal("0")) + (female_total or Decimal("0"))
        elif categories:
            total_value = sum_category_payload(categories)
        else:
            totals = [pick_total(row) for row in section["rows"]]
            totals = [total for total in totals if total is not None]
            total_value = totals[-1] if totals else Decimal("0")

    payload = {
        "total": decimal_to_json_number(total_value or Decimal("0")),
    }
    if male_total is not None:
        payload["male"] = decimal_to_json_number(male_total)
    if female_total is not None:
        payload["female"] = decimal_to_json_number(female_total)
    if disaggregate_matrix:
        payload["disaggregates"] = disaggregate_matrix
    if categories:
        payload["categories"] = categories

    return payload, disaggregations


class Command(BaseCommand):
    help = "Import a structured quarterly reporting workbook, create missing indicators, assign targets, and load aggregates."

    def add_arguments(self, parser):
        parser.add_argument("workbook", help="Path to the .xlsx workbook")
        parser.add_argument("--project-id", type=int, required=True, help="Project to assign indicators and aggregates to")
        parser.add_argument("--coordinator-id", type=int, help="Optional coordinator organization for rollup aggregates")
        parser.add_argument("--period-start", required=True, help="Quarter start date, for example 2025-10-01")
        parser.add_argument("--period-end", required=True, help="Quarter end date, for example 2025-12-31")
        parser.add_argument("--category", default="ncd", help="Indicator category to use for new indicators")
        parser.add_argument("--dry-run", action="store_true", help="Parse and print what would be imported without saving")

    def handle(self, *args, **options):
        workbook_path = Path(options["workbook"]).expanduser()
        if not workbook_path.exists():
            raise CommandError(f"Workbook not found: {workbook_path}")

        project = Project.objects.filter(id=options["project_id"]).first()
        if not project:
            raise CommandError(f"Project not found: {options['project_id']}")

        coordinator = None
        if options.get("coordinator_id"):
            coordinator = Organization.objects.filter(id=options["coordinator_id"]).first()
            if not coordinator:
                raise CommandError(f"Coordinator organization not found: {options['coordinator_id']}")

        workbook = load_workbook(workbook_path, data_only=True, read_only=True)
        formula_workbook = load_workbook(workbook_path, data_only=False, read_only=True)
        matrix_sheet_name = find_matrix_sheet_name(list(workbook.sheetnames))
        if not matrix_sheet_name:
            raise CommandError("Workbook is missing the Indicator matrix sheet.")

        matrix_assignments = parse_matrix_sheet(workbook[matrix_sheet_name])
        organization_resolver = OrganizationResolver()
        indicator_resolver = IndicatorResolver(project=project)

        sheet_payloads = []
        for sheet_name in workbook.sheetnames:
            if is_skipped_sheet(sheet_name):
                continue

            organization = organization_resolver.resolve(sheet_name)
            if not organization:
                self.stderr.write(f"Skipping sheet '{sheet_name}': organization could not be resolved.")
                continue

            formula_sheet = formula_workbook[sheet_name] if sheet_name in formula_workbook.sheetnames else None
            sections = parse_sections(workbook[sheet_name], formula_sheet)
            parsed_sections = []
            age_band_by_column = get_age_band_mapping(organization, coordinator)
            for section in sections:
                value, disaggregations = extract_section_value(
                    section,
                    age_band_by_column=age_band_by_column,
                )
                parsed_sections.append(
                    {
                        "title": section["title"],
                        "index": section["index"],
                        "value": value,
                        "disaggregations": disaggregations,
                    }
                )
            sheet_payloads.append(
                {
                    "sheet_name": sheet_name,
                    "organization": organization,
                    "sections": parsed_sections,
                }
            )

        if options["dry_run"]:
            self.stdout.write(self.style.WARNING("Dry run only. No database changes were written."))
            self.stdout.write(f"Workbook: {workbook_path}")
            self.stdout.write(f"Project: {project.id} - {project.name}")
            self.stdout.write(f"Organizations resolved: {len(sheet_payloads)}")
            self.stdout.write(f"Matrix assignments found: {len(matrix_assignments)}")
            for payload in sheet_payloads[:5]:
                self.stdout.write(f"- {payload['sheet_name']} -> {payload['organization'].name}: {len(payload['sections'])} sections")
            return

        created_indicators = 0
        updated_targets = 0
        imported_aggregates = 0
        coordinator_rollups = {}
        coordinator_source_sheet_seen = False

        with transaction.atomic():
            for payload in sheet_payloads:
                organization = payload["organization"]
                if coordinator and organization.id == coordinator.id:
                    coordinator_source_sheet_seen = True
                project.organizations.add(organization)

                for section in payload["sections"]:
                    title = section["title"]
                    if not str(title or "").strip():
                        continue
                    indicator = indicator_resolver.resolve(title, section["index"])
                    desired_sub_labels = build_ordered_sub_labels(
                        section["disaggregations"],
                        section["value"],
                    )
                    desired_config = build_disaggregation_config(
                        section["value"],
                        desired_sub_labels,
                    )
                    if not indicator:
                        code_seed = f"{project.code}-{section['index']}-{title}"
                        indicator = Indicator.objects.create(
                            name=title,
                            code=unique_indicator_code(code_seed),
                            type="number",
                            category=options["category"],
                            unit="people",
                            sub_labels=desired_sub_labels,
                        )
                        indicator_resolver.remember(indicator)
                        created_indicators += 1
                    else:
                        indicator_updates = []
                        if list(indicator.sub_labels or []) != desired_sub_labels:
                            indicator.sub_labels = desired_sub_labels
                            indicator_updates.append("sub_labels")

                        if indicator_updates:
                            indicator.save(update_fields=indicator_updates)

                    indicator.organizations.add(organization)

                    project_indicator, _ = ensure_project_indicator(
                        project=project,
                        indicator=indicator,
                    )

                    assignment_bundle = matrix_assignments.get(canonical_indicator_name(title))
                    if assignment_bundle:
                        for assignment in assignment_bundle["assignments"]:
                            assigned_organization = organization_resolver.resolve(assignment["organization_name"])
                            if not assigned_organization or assigned_organization.id != organization.id:
                                continue
                            project.organizations.add(assigned_organization)
                            indicator.organizations.add(assigned_organization)

                    aggregate, created = Aggregate.objects.get_or_create(
                        indicator=indicator,
                        project=project,
                        organization=organization,
                        period_start=options["period_start"],
                        period_end=options["period_end"],
                        defaults={"value": section["value"]},
                    )
                    if not created:
                        if aggregate.value != section["value"]:
                            aggregate.value = section["value"]
                            aggregate.save(update_fields=["value"])
                    imported_aggregates += 1

                    if coordinator and organization.id != coordinator.id:
                        project.organizations.add(coordinator)
                        indicator.organizations.add(coordinator)
                        key = indicator.id
                        if key in coordinator_rollups:
                            coordinator_rollups[key]["value"] = merge_json_values(
                                coordinator_rollups[key]["value"],
                                section["value"],
                            )
                        else:
                            coordinator_rollups[key] = {
                                "indicator": indicator,
                                "value": section["value"],
                            }

            if coordinator and not coordinator_source_sheet_seen:
                for rollup in coordinator_rollups.values():
                    Aggregate.objects.update_or_create(
                        indicator=rollup["indicator"],
                        project=project,
                        organization=coordinator,
                        period_start=options["period_start"],
                        period_end=options["period_end"],
                        defaults={"value": rollup["value"]},
                    )

        self.stdout.write(self.style.SUCCESS("Workbook import complete."))
        self.stdout.write(f"Created indicators: {created_indicators}")
        self.stdout.write(f"Updated org targets: {updated_targets}")
        self.stdout.write(f"Imported aggregates: {imported_aggregates}")
        if coordinator:
            if coordinator_source_sheet_seen:
                self.stdout.write(
                    self.style.WARNING(
                        "Coordinator rollups skipped: coordinator has its own source sheet, "
                        "so writing rollups would overwrite organization values."
                    )
                )
            else:
                self.stdout.write(f"Coordinator rollups: {len(coordinator_rollups)}")
