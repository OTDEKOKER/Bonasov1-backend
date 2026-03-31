import re
from decimal import Decimal
from pathlib import Path

from django.db import transaction
from django.core.files.base import ContentFile
from django.http import Http404, HttpResponse
from django.utils import timezone
from rest_framework import mixins, serializers, status, viewsets
from rest_framework.decorators import action
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from aggregates.models import Aggregate
from indicators.models import Indicator
from organizations.models import Organization
from projects.models import Project, ProjectIndicator, ProjectIndicatorOrganizationTarget
from openpyxl import Workbook
from uploads.management.commands.import_reporting_workbook import (
    IndicatorResolver,
    OrganizationResolver,
    build_disaggregation_config,
    build_ordered_sub_labels,
    canonical_indicator_name,
    extract_section_value,
    find_matrix_sheet_name,
    get_age_band_mapping,
    is_skipped_sheet,
    merge_disaggregation_configs,
    merge_json_values,
    parse_matrix_sheet,
    parse_sections,
    unique_indicator_code,
)

from .models import ImportJob, Upload, WorkbookExportJob, WorkbookTemplate
from .serializers import CreateMissingIndicatorsSerializer
from .views import _resolve_assignment_organization, _sanitize_indicator_code, _to_decimal

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None


SESSION_KEY = "report_workbook_session"
ANALYSIS_KEY = "report_workbook_analysis"


def _json_number(value):
    if isinstance(value, Decimal):
        return int(value) if value == value.to_integral_value() else float(value)
    return value


def _normalize_json(value):
    if isinstance(value, Decimal):
        return _json_number(value)
    if isinstance(value, dict):
        return {key: _normalize_json(item) for key, item in value.items()}
    if isinstance(value, (list, tuple)):
        return [_normalize_json(item) for item in value]
    return value


def _get_payload(job: ImportJob, payload_type: str) -> dict:
    for item in list(job.errors or []):
        if isinstance(item, dict) and item.get("type") == payload_type:
            return dict(item.get("data") or {})
    return {}


def _set_payload(job: ImportJob, payload_type: str, data: dict) -> None:
    entries = []
    replaced = False
    for item in list(job.errors or []):
        if isinstance(item, dict) and item.get("type") == payload_type:
            if not replaced:
                entries.append({"type": payload_type, "data": _normalize_json(data)})
                replaced = True
            continue
        entries.append(item)
    if not replaced:
        entries.append({"type": payload_type, "data": _normalize_json(data)})
    job.errors = entries


def _build_summary(**overrides) -> dict:
    summary = {
        "sheets_scanned": 0,
        "sheets_skipped": 0,
        "rows_imported": 0,
        "rows_skipped": 0,
        "assignments_detected": 0,
        "target_cells_detected": 0,
        "warnings": 0,
        "errors": 0,
    }
    summary.update(overrides)
    return summary


def _parse_reporting_period(period_label: str | None) -> tuple[str, str, int]:
    text = str(period_label or "").strip()
    match = re.search(r"\bQ([1-4])\s+(\d{4})\b", text, re.IGNORECASE)
    if not match:
        raise serializers.ValidationError(
            {"reporting_period": "Expected a period like 'Q3 2025 (Oct-Dec)'."}
        )
    quarter = int(match.group(1))
    year = int(match.group(2))
    if quarter == 1:
        return f"{year}-04-01", f"{year}-06-30", 4
    if quarter == 2:
        return f"{year}-07-01", f"{year}-09-30", 4
    if quarter == 3:
        return f"{year}-10-01", f"{year}-12-31", 4
    return f"{year + 1}-01-01", f"{year + 1}-03-31", 4


def _get_workbook(upload: Upload):
    if load_workbook is None:
        raise serializers.ValidationError({"detail": "Workbook support is not available."})
    path = Path(upload.file.path)
    if not path.exists():
        raise Http404("Uploaded workbook file could not be found.")
    return load_workbook(path, data_only=True)


def _update_status(job: ImportJob, value: str) -> None:
    session = _get_payload(job, SESSION_KEY)
    session["status"] = value
    _set_payload(job, SESSION_KEY, session)


def _session_status(job: ImportJob) -> str:
    return _get_payload(job, SESSION_KEY).get("status") or "uploaded"


def _serialize_session(job: ImportJob) -> dict:
    session = _get_payload(job, SESSION_KEY)
    analysis = _get_payload(job, ANALYSIS_KEY)
    project = Project.objects.filter(id=session.get("project")).first() if session.get("project") else None
    organization = Organization.objects.filter(id=session.get("organization")).first() if session.get("organization") else None
    template = WorkbookTemplate.objects.filter(id=session.get("template_id")).first() if session.get("template_id") else None
    return {
        "id": job.id,
        "file_name": job.upload.file.name.split("/")[-1] if job.upload.file else job.upload.name,
        "status": _session_status(job),
        "project": project.id if project else None,
        "project_name": project.name if project else None,
        "organization": organization.id if organization else None,
        "organization_name": organization.name if organization else None,
        "reporting_period": session.get("reporting_period"),
        "template_id": session.get("template_id"),
        "template_name": template.name if template else None,
        "created_at": job.created_at,
        "updated_at": job.completed_at or job.started_at or job.created_at,
        "financial_year_start_month": analysis.get("financial_year_start_month"),
        "summary": analysis.get("summary"),
        "assignments": analysis.get("assignments", []),
        "missing_indicators": analysis.get("missing_indicators", []),
        "sheets": analysis.get("sheets", []),
        "issues": analysis.get("issues", []),
    }


class ReportWorkbookImportCreateSerializer(serializers.Serializer):
    file = serializers.FileField()
    project = serializers.IntegerField(required=False)
    organization = serializers.IntegerField(required=False)
    reporting_period = serializers.CharField(required=False, allow_blank=True)
    template_id = serializers.IntegerField(required=False)
    notes = serializers.CharField(required=False, allow_blank=True)


class ReportWorkbookConfirmSerializer(serializers.Serializer):
    import_mode = serializers.ChoiceField(choices=["append", "replace_period"], required=False, default="append")
    overwrite_existing = serializers.BooleanField(required=False, default=False)
    apply_indicator_assignments = serializers.BooleanField(required=False, default=True)
    sync_project_indicator_links = serializers.BooleanField(required=False, default=True)
    create_missing_indicators = serializers.BooleanField(required=False, default=False)


class ReportWorkbookExportCreateSerializer(serializers.Serializer):
    project = serializers.IntegerField()
    reporting_period = serializers.CharField()
    scope = serializers.ChoiceField(
        choices=["single_organization", "coordinator", "all_organizations", "consolidated"]
    )
    financial_year_start_month = serializers.IntegerField(required=False, min_value=1, max_value=12, default=4)
    organization_ids = serializers.ListField(child=serializers.IntegerField(), required=False)
    coordinator_id = serializers.IntegerField(required=False)
    template_id = serializers.IntegerField(required=False)
    preserve_formatting = serializers.BooleanField(required=False, default=True)
    include_validation_summary = serializers.BooleanField(required=False, default=True)

    def validate(self, attrs):
        scope = attrs.get("scope")
        organization_ids = attrs.get("organization_ids") or []
        coordinator_id = attrs.get("coordinator_id")
        if scope == "single_organization" and not organization_ids:
            raise serializers.ValidationError(
                {"organization_ids": "Provide at least one organization id for single_organization exports."}
            )
        if scope == "coordinator" and not coordinator_id:
            raise serializers.ValidationError(
                {"coordinator_id": "Provide coordinator_id for coordinator exports."}
            )
        return attrs


def _build_assignment_preview(indicator_name: str, assignment: dict, source_sheet: str) -> dict:
    annual_target = sum((_to_decimal(assignment.get(f"q{index}_target")) for index in range(1, 5)), Decimal("0"))
    return {
        "organization_id": assignment.get("organization_id"),
        "organization_name": assignment.get("organization_name") or "Unknown",
        "coordinator_id": None,
        "coordinator_name": None,
        "indicator_id": assignment.get("indicator_id"),
        "indicator_name": indicator_name,
        "financial_year_label": None,
        "targets_by_quarter": {
            "Q1": _json_number(assignment.get("q1_target")),
            "Q2": _json_number(assignment.get("q2_target")),
            "Q3": _json_number(assignment.get("q3_target")),
            "Q4": _json_number(assignment.get("q4_target")),
        },
        "annual_target": _json_number(annual_target),
        "source_sheet": source_sheet,
        "source_row_ref": None,
        "assignment_confidence": 0.95,
    }


def _build_missing_indicator_candidate(
    title: str,
    source_sheet: str,
    section_index: str,
    assignments: list[dict],
    disaggregations: set[str],
    value: dict,
) -> dict:
    sub_labels = build_ordered_sub_labels(disaggregations, value)
    config = build_disaggregation_config(value, sub_labels)
    return {
        "temp_key": f"{section_index or 'x'}-{_sanitize_indicator_code(title)}",
        "indicator_name": title,
        "suggested_code": _sanitize_indicator_code(title),
        "category": "ncd",
        "type": "multi_int" if config.get("enabled") else "number",
        "unit": "people",
        "sub_labels": sub_labels,
        "aggregate_disaggregation_config": config or None,
        "disaggregation_preset_keys": [],
        "source_sheet": source_sheet,
        "source_row_ref": None,
        "assigned_organizations": [
            {
                "organization_id": assignment.get("organization_id"),
                "organization_name": assignment.get("organization_name") or "Unknown",
                "q1_target": _json_number(assignment.get("q1_target")),
                "q2_target": _json_number(assignment.get("q2_target")),
                "q3_target": _json_number(assignment.get("q3_target")),
                "q4_target": _json_number(assignment.get("q4_target")),
            }
            for assignment in assignments
        ],
    }


def _analyze_workbook(job: ImportJob) -> dict:
    workbook = _get_workbook(job.upload)
    session = _get_payload(job, SESSION_KEY)
    project = Project.objects.filter(id=session.get("project")).first() if session.get("project") else None
    indicator_resolver = IndicatorResolver(project=project) if project else IndicatorResolver()
    organization_resolver = OrganizationResolver()
    issues = []
    assignments = []
    missing = []
    sheets = []
    names = list(workbook.sheetnames)
    matrix_sheet_name = find_matrix_sheet_name(names)
    if not matrix_sheet_name:
        issues.append(
            {
                "severity": "error",
                "code": "missing_indicator_matrix",
                "message": "Workbook is missing the Indicator matrix sheet.",
                "sheet_name": None,
                "cell_ref": None,
                "details": None,
            }
        )
        return {
            "summary": _build_summary(sheets_scanned=len(names), errors=1),
            "assignments": [],
            "missing_indicators": [],
            "sheets": [],
            "issues": issues,
            "financial_year_start_month": 4,
        }

    matrix_assignments = parse_matrix_sheet(workbook[matrix_sheet_name])
    row_count = 0
    skipped_count = 0
    for index, sheet_name in enumerate(names):
        ws = workbook[sheet_name]
        title_blocks = []
        metadata_cells = {}
        for row_number in range(1, min(ws.max_row, 8) + 1):
            for column in ("A", "B", "C", "D", "E", "F"):
                value = ws[f"{column}{row_number}"].value
                if value not in (None, "") and len(metadata_cells) < 12:
                    metadata_cells[f"{column}{row_number}"] = str(value).strip()
                    title_blocks.append(str(value).strip())

        role = "unknown"
        detected_org = None
        sheet_issues = []
        table_previews = []
        if sheet_name == matrix_sheet_name:
            role = "indicator_matrix"
            for indicator_key, bundle in matrix_assignments.items():
                for assignment in bundle.get("assignments", []):
                    assignments.append(_build_assignment_preview(bundle.get("title") or indicator_key, assignment, sheet_name))
        elif is_skipped_sheet(sheet_name):
            role = "summary"
            skipped_count += 1
        else:
            organization = organization_resolver.resolve(sheet_name)
            if organization:
                role = "organization_report"
                detected_org = organization.name
                sections = parse_sections(ws)
                row_count += len(sections)
                age_band_by_column = get_age_band_mapping(
                    organization,
                    Organization.objects.filter(id=session.get("organization")).first()
                    if session.get("organization")
                    else None,
                )
                table_previews.append(
                    {
                        "title": sheet_name,
                        "start_cell": "A1",
                        "end_cell": f"AA{ws.max_row}",
                        "header_rows": [1, 2, 3],
                        "row_label_columns": ["B", "C", "E", "F"],
                        "column_headers": list(age_band_by_column.values()),
                        "row_labels": [section["title"] for section in sections[:20]],
                        "totals_rows": [],
                        "numeric_cells_detected": len(sections),
                    }
                )
                for section in sections:
                    value, disaggregations = extract_section_value(section, age_band_by_column=age_band_by_column)
                    indicator = indicator_resolver.resolve(section["title"], section["index"])
                    matched_assignments = list(
                        matrix_assignments.get(canonical_indicator_name(section["title"]), {}).get("assignments", [])
                    )
                    if not indicator:
                        if not matched_assignments:
                            matched_assignments = [
                                {
                                    "organization_id": organization.id,
                                    "organization_name": organization.name,
                                    "q1_target": None,
                                    "q2_target": None,
                                    "q3_target": None,
                                    "q4_target": None,
                                }
                            ]
                        missing.append(
                            _build_missing_indicator_candidate(
                                section["title"],
                                sheet_name,
                                section["index"],
                                matched_assignments,
                                disaggregations,
                                value,
                            )
                        )
            else:
                skipped_count += 1
                sheet_issues.append(
                    {
                        "severity": "warning",
                        "code": "organization_unresolved",
                        "message": f"Could not resolve organization from worksheet name '{sheet_name}'.",
                        "sheet_name": sheet_name,
                        "cell_ref": None,
                        "details": None,
                    }
                )
                issues.extend(sheet_issues)

        sheets.append(
            {
                "id": index + 1,
                "sheet_name": sheet_name,
                "sheet_index": index,
                "sheet_role": role,
                "detected_indicator": None,
                "detected_organization": detected_org,
                "detected_project": project.name if project else None,
                "detected_reporting_period": session.get("reporting_period"),
                "merged_ranges": [str(cell_range) for cell_range in list(ws.merged_cells.ranges)[:20]],
                "title_blocks": title_blocks[:10],
                "metadata_cells": metadata_cells,
                "table_previews": table_previews,
                "assignment_previews": [item for item in assignments if item["source_sheet"] == sheet_name],
                "template_match": None,
                "issues": sheet_issues,
            }
        )

    deduped_missing = []
    seen = set()
    for item in missing:
        key = (item["source_sheet"], canonical_indicator_name(item["indicator_name"]))
        if key in seen:
            continue
        seen.add(key)
        deduped_missing.append(item)

    summary = _build_summary(
        sheets_scanned=len(names),
        sheets_skipped=skipped_count,
        rows_imported=row_count,
        assignments_detected=len(assignments),
        target_cells_detected=len(assignments) * 4,
        warnings=sum(1 for issue in issues if issue["severity"] == "warning"),
        errors=sum(1 for issue in issues if issue["severity"] == "error"),
    )
    return {
        "summary": summary,
        "assignments": assignments,
        "missing_indicators": deduped_missing,
        "sheets": sheets,
        "issues": issues,
        "financial_year_start_month": 4,
    }


def _upsert_template_from_analysis(job: ImportJob, analysis: dict) -> WorkbookTemplate:
    session = _get_payload(job, SESSION_KEY)
    upload_name = Path(job.upload.name or "").stem or Path(job.upload.file.name).stem
    header_values = []
    row_labels = []
    column_labels = []
    for sheet in analysis.get("sheets", []):
        header_values.extend(list((sheet.get("metadata_cells") or {}).values()))
        for preview in sheet.get("table_previews", []):
            row_labels.extend(preview.get("row_labels") or [])
            column_labels.extend(preview.get("column_headers") or [])

    def _unique(values):
        seen = set()
        ordered = []
        for value in values:
            text = str(value or "").strip()
            if not text or text in seen:
                continue
            seen.add(text)
            ordered.append(text)
        return ordered

    template, _ = WorkbookTemplate.objects.update_or_create(
        source_upload=job.upload,
        defaults={
            "name": upload_name,
            "workbook_family": "report_workbook",
            "report_category": "aggregate_reporting",
            "version": 1,
            "is_active": True,
            "expected_headers": _unique(header_values)[:50],
            "row_labels": _unique(row_labels)[:200],
            "column_labels": _unique(column_labels)[:100],
            "created_by": job.created_by,
        },
    )
    return template


def _apply_missing_indicators(job: ImportJob, payload: dict, user) -> None:
    session = _get_payload(job, SESSION_KEY)
    project = Project.objects.filter(id=session.get("project")).first() if session.get("project") else None
    assign_to_project = payload.get("assign_to_project", True)
    create_targets = payload.get("create_targets", True)

    with transaction.atomic():
        for item in payload["indicators"]:
            name = item["name"].strip()
            code = _sanitize_indicator_code(item["code"])
            indicator = Indicator.objects.filter(code__iexact=code).first() or Indicator.objects.filter(name__iexact=name).first()
            if not indicator:
                indicator = Indicator.objects.create(
                    name=name,
                    code=unique_indicator_code(code),
                    type=item.get("type") or "number",
                    category=item.get("category") or "ncd",
                    unit=item.get("unit", ""),
                    sub_labels=item.get("sub_labels") or [],
                    aggregate_disaggregation_config=item.get("aggregate_disaggregation_config") or {},
                    created_by=user,
                )
            else:
                updated_fields = []
                if item.get("sub_labels") and not list(indicator.sub_labels or []):
                    indicator.sub_labels = item.get("sub_labels") or []
                    updated_fields.append("sub_labels")
                merged_config = merge_disaggregation_configs(
                    indicator.aggregate_disaggregation_config,
                    item.get("aggregate_disaggregation_config") or {},
                )
                if merged_config != dict(indicator.aggregate_disaggregation_config or {}):
                    indicator.aggregate_disaggregation_config = merged_config
                    updated_fields.append("aggregate_disaggregation_config")
                if updated_fields:
                    indicator.save(update_fields=updated_fields)

            resolved_organizations = {}
            for org_id in item.get("organizations", []):
                organization = Organization.objects.filter(id=org_id).first()
                if organization:
                    resolved_organizations[str(organization.id)] = organization

            project_indicator = None
            if assign_to_project and project:
                project_indicator, _ = ProjectIndicator.objects.get_or_create(project=project, indicator=indicator)

            for assignment in item.get("assignments", []):
                organization = _resolve_assignment_organization(assignment)
                if not organization:
                    continue
                resolved_organizations[str(organization.id)] = organization
                if create_targets and project_indicator:
                    org_target, _ = ProjectIndicatorOrganizationTarget.objects.get_or_create(
                        project_indicator=project_indicator,
                        organization=organization,
                    )
                    org_target.q1_target = _to_decimal(assignment.get("q1_target"))
                    org_target.q2_target = _to_decimal(assignment.get("q2_target"))
                    org_target.q3_target = _to_decimal(assignment.get("q3_target"))
                    org_target.q4_target = _to_decimal(assignment.get("q4_target"))
                    org_target.save()

            if resolved_organizations:
                indicator.organizations.add(*resolved_organizations.values())
                if project:
                    project.organizations.add(*resolved_organizations.values())


def _confirm_import(job: ImportJob, payload: dict, user) -> dict:
    session = _get_payload(job, SESSION_KEY)
    project = Project.objects.filter(id=session.get("project")).first() if session.get("project") else None
    if not project:
        raise serializers.ValidationError({"project": "A project must be selected before confirming import."})

    period_start, period_end, _ = _parse_reporting_period(session.get("reporting_period"))
    workbook = _get_workbook(job.upload)
    matrix_sheet_name = find_matrix_sheet_name(list(workbook.sheetnames))
    if not matrix_sheet_name:
        raise serializers.ValidationError({"detail": "Workbook is missing the Indicator matrix sheet."})

    if payload.get("create_missing_indicators"):
        analysis = _get_payload(job, ANALYSIS_KEY)
        missing_payload = {
            "assign_to_project": payload.get("sync_project_indicator_links", True),
            "create_targets": payload.get("apply_indicator_assignments", True),
            "indicators": [
                {
                    "temp_key": item["temp_key"],
                    "name": item["indicator_name"],
                    "code": item.get("suggested_code") or item["temp_key"],
                    "type": item.get("type") or "number",
                    "category": item.get("category") or "ncd",
                    "unit": item.get("unit") or "people",
                    "sub_labels": item.get("sub_labels") or [],
                    "aggregate_disaggregation_config": item.get("aggregate_disaggregation_config") or {},
                    "organizations": [org["organization_id"] for org in item.get("assigned_organizations", []) if org.get("organization_id")],
                    "assignments": [
                        {
                            "organization_id": org.get("organization_id"),
                            "organization_name": org.get("organization_name"),
                            "q1_target": org.get("q1_target"),
                            "q2_target": org.get("q2_target"),
                            "q3_target": org.get("q3_target"),
                            "q4_target": org.get("q4_target"),
                        }
                        for org in item.get("assigned_organizations", [])
                    ],
                }
                for item in analysis.get("missing_indicators", [])
            ],
        }
        if missing_payload["indicators"]:
            _apply_missing_indicators(job, missing_payload, user)

    coordinator = Organization.objects.filter(id=session.get("organization")).first() if session.get("organization") else None
    matrix_assignments = parse_matrix_sheet(workbook[matrix_sheet_name])
    organization_resolver = OrganizationResolver()
    indicator_resolver = IndicatorResolver(project=project)
    replace_period = payload.get("import_mode") == "replace_period"
    overwrite_existing = bool(payload.get("overwrite_existing"))
    apply_assignments = payload.get("apply_indicator_assignments", True)

    organization_payloads = []
    for sheet_name in workbook.sheetnames:
        if is_skipped_sheet(sheet_name):
            continue
        organization = organization_resolver.resolve(sheet_name)
        if not organization:
            continue
        age_band_by_column = get_age_band_mapping(organization, coordinator)
        organization_payloads.append(
            {
                "organization": organization,
                "sections": [
                    {
                        "title": section["title"],
                        "index": section["index"],
                        "value": extract_section_value(section, age_band_by_column=age_band_by_column)[0],
                        "disaggregations": extract_section_value(section, age_band_by_column=age_band_by_column)[1],
                    }
                    for section in parse_sections(workbook[sheet_name])
                ],
            }
        )

    scope_ids = [item["organization"].id for item in organization_payloads]
    if coordinator:
        scope_ids.append(coordinator.id)
    if replace_period and scope_ids:
        Aggregate.objects.filter(
            project=project,
            organization_id__in=scope_ids,
            period_start=period_start,
            period_end=period_end,
        ).delete()

    imported_rows = 0
    skipped_rows = 0
    issues = []
    coordinator_rollups = {}

    with transaction.atomic():
        for payload_item in organization_payloads:
            organization = payload_item["organization"]
            project.organizations.add(organization)
            for section in payload_item["sections"]:
                indicator = indicator_resolver.resolve(section["title"], section["index"])
                if not indicator:
                    skipped_rows += 1
                    issues.append(
                        {
                            "severity": "warning",
                            "code": "indicator_missing",
                            "message": f"Skipped '{section['title']}' for {organization.name} because no indicator could be resolved.",
                            "sheet_name": organization.name,
                            "cell_ref": None,
                            "details": {"section_index": section["index"]},
                        }
                    )
                    continue

                sub_labels = build_ordered_sub_labels(section["disaggregations"], section["value"])
                config = build_disaggregation_config(section["value"], sub_labels)
                updated_fields = []
                if sub_labels and not list(indicator.sub_labels or []):
                    indicator.sub_labels = sub_labels
                    updated_fields.append("sub_labels")
                merged_config = merge_disaggregation_configs(indicator.aggregate_disaggregation_config, config)
                if merged_config != dict(indicator.aggregate_disaggregation_config or {}):
                    indicator.aggregate_disaggregation_config = merged_config
                    updated_fields.append("aggregate_disaggregation_config")
                if updated_fields:
                    indicator.save(update_fields=updated_fields)

                indicator.organizations.add(organization)
                project_indicator, _ = ProjectIndicator.objects.get_or_create(project=project, indicator=indicator)
                if apply_assignments:
                    bundle = matrix_assignments.get(canonical_indicator_name(section["title"]))
                    if bundle:
                        for assignment in bundle["assignments"]:
                            assigned_org = organization_resolver.resolve(assignment["organization_name"])
                            if not assigned_org or assigned_org.id != organization.id:
                                continue
                            project.organizations.add(assigned_org)
                            indicator.organizations.add(assigned_org)
                            org_target, _ = ProjectIndicatorOrganizationTarget.objects.get_or_create(
                                project_indicator=project_indicator,
                                organization=assigned_org,
                            )
                            org_target.q1_target = assignment["q1_target"]
                            org_target.q2_target = assignment["q2_target"]
                            org_target.q3_target = assignment["q3_target"]
                            org_target.q4_target = assignment["q4_target"]
                            org_target.save()

                aggregate, created = Aggregate.objects.get_or_create(
                    indicator=indicator,
                    project=project,
                    organization=organization,
                    period_start=period_start,
                    period_end=period_end,
                    defaults={"value": section["value"], "status": Aggregate.STATUS_PENDING, "created_by": user},
                )
                if not created:
                    aggregate.value = section["value"] if overwrite_existing or replace_period else merge_json_values(aggregate.value, section["value"])
                    aggregate.status = Aggregate.STATUS_PENDING
                    if not aggregate.created_by:
                        aggregate.created_by = user
                    aggregate.save(update_fields=["value", "status", "created_by", "updated_at"])
                imported_rows += 1

                if coordinator:
                    current = coordinator_rollups.get(indicator.id)
                    coordinator_rollups[indicator.id] = {
                        "indicator": indicator,
                        "value": section["value"] if current is None else merge_json_values(current["value"], section["value"]),
                    }

        if coordinator:
            for item in coordinator_rollups.values():
                Aggregate.objects.update_or_create(
                    indicator=item["indicator"],
                    project=project,
                    organization=coordinator,
                    period_start=period_start,
                    period_end=period_end,
                    defaults={"value": item["value"], "status": Aggregate.STATUS_PENDING, "created_by": user},
                )

    return {
        "summary": _build_summary(
            sheets_scanned=len(organization_payloads) + 1,
            sheets_skipped=max(0, len(workbook.sheetnames) - (len(organization_payloads) + 1)),
            rows_imported=imported_rows,
            rows_skipped=skipped_rows,
            assignments_detected=sum(len(bundle.get("assignments", [])) for bundle in matrix_assignments.values()),
            target_cells_detected=sum(len(bundle.get("assignments", [])) for bundle in matrix_assignments.values()) * 4,
            warnings=len(issues),
            errors=0,
        ),
        "issues": issues,
    }


def _serialize_template(template: WorkbookTemplate) -> dict:
    return {
        "id": template.id,
        "name": template.name,
        "workbook_family": template.workbook_family,
        "report_category": template.report_category,
        "version": template.version,
        "is_active": template.is_active,
        "expected_headers": template.expected_headers or [],
        "row_labels": template.row_labels or [],
        "column_labels": template.column_labels or [],
    }


def _build_export_workbook(job: WorkbookExportJob, include_validation_summary: bool = True) -> Upload:
    period_start, period_end, _ = _parse_reporting_period(job.reporting_period)
    aggregates = Aggregate.objects.filter(project=job.project, period_start=period_start, period_end=period_end).select_related(
        "indicator", "organization"
    )
    if job.scope == "single_organization" and job.organization_id:
        aggregates = aggregates.filter(organization_id=job.organization_id)
    elif job.scope == "coordinator" and job.coordinator_id:
        aggregates = aggregates.filter(organization_id=job.coordinator_id)
    elif job.scope == "all_organizations":
        organization_ids = [job.organization_id] if job.organization_id else []
        if organization_ids:
            aggregates = aggregates.filter(organization_id__in=organization_ids)

    workbook = Workbook()
    summary_ws = workbook.active
    summary_ws.title = "Summary"
    summary_ws.append(["Project", job.project.name])
    summary_ws.append(["Reporting period", job.reporting_period])
    summary_ws.append(["Scope", job.scope])
    if job.organization:
        summary_ws.append(["Organization", job.organization.name])
    if job.coordinator:
        summary_ws.append(["Coordinator", job.coordinator.name])
    summary_ws.append([])
    summary_ws.append(["Indicator Code", "Indicator", "Organization", "Total", "Value JSON"])

    organization_sheets = {}
    row_count = 0
    for aggregate in aggregates.order_by("organization__name", "indicator__code"):
        total = aggregate.value.get("total") if isinstance(aggregate.value, dict) else aggregate.value
        summary_ws.append(
            [
                aggregate.indicator.code,
                aggregate.indicator.name,
                aggregate.organization.name,
                total,
                str(aggregate.value),
            ]
        )
        row_count += 1
        sheet_name = re.sub(r"[\[\]\*:/\\\?]", "-", aggregate.organization.name)[:31] or f"Org-{aggregate.organization_id}"
        ws = organization_sheets.get(sheet_name)
        if ws is None:
            ws = workbook.create_sheet(title=sheet_name)
            ws.append(["Indicator Code", "Indicator", "Period Start", "Period End", "Total", "Value JSON"])
            organization_sheets[sheet_name] = ws
        ws.append(
            [
                aggregate.indicator.code,
                aggregate.indicator.name,
                str(aggregate.period_start),
                str(aggregate.period_end),
                total,
                str(aggregate.value),
            ]
        )

    if include_validation_summary:
        validation_ws = workbook.create_sheet(title="Validation")
        validation_ws.append(["Metric", "Value"])
        validation_ws.append(["Rows exported", row_count])
        validation_ws.append(["Generated at", timezone.now().isoformat()])
        validation_ws.append(["Template", job.template.name if job.template else "None"])

    from io import BytesIO

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    file_name = f"{job.project.code or job.project.id}-{job.reporting_period.replace(' ', '_')}-{job.scope}.xlsx"
    upload = Upload(
        name=file_name,
        description=f"Generated workbook export for {job.project.name} {job.reporting_period}",
        organization=job.organization or job.coordinator,
        created_by=job.created_by,
        content_type="report_workbook_export",
        object_id=job.id,
    )
    upload.file.save(file_name, ContentFile(buffer.getvalue()), save=True)
    return upload


def _serialize_export_job(job: WorkbookExportJob, request=None) -> dict:
    download_url = None
    if job.generated_upload and job.generated_upload.file:
        if request:
            download_url = request.build_absolute_uri(f"/api/report-workbooks/exports/{job.id}/download/")
        else:
            download_url = f"/api/report-workbooks/exports/{job.id}/download/"
    return {
        "id": job.id,
        "status": job.status,
        "file_name": job.generated_upload.file.name.split("/")[-1] if job.generated_upload and job.generated_upload.file else None,
        "download_url": download_url,
        "scope": job.scope,
        "reporting_period": job.reporting_period,
        "project": job.project_id,
        "created_at": job.created_at,
        "completed_at": job.completed_at,
        "errors": job.errors or [],
    }


class ReportWorkbookImportViewSet(mixins.CreateModelMixin, mixins.ListModelMixin, mixins.RetrieveModelMixin, viewsets.GenericViewSet):
    queryset = ImportJob.objects.select_related("upload", "created_by").all()
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    def get_queryset(self):
        user = self.request.user
        if user.is_superuser or user.is_staff or getattr(user, "role", "") == "admin":
            return self.queryset
        if getattr(user, "organization_id", None):
            return self.queryset.filter(upload__organization_id=user.organization_id)
        return self.queryset.filter(created_by=user)

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        page = self.paginate_queryset(queryset)
        items = page if page is not None else queryset
        data = [_serialize_session(job) for job in items]
        if page is not None:
            return self.get_paginated_response(data)
        return Response(data)

    def retrieve(self, request, *args, **kwargs):
        return Response(_serialize_session(self.get_object()))

    def create(self, request, *args, **kwargs):
        serializer = ReportWorkbookImportCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        payload = serializer.validated_data

        upload = Upload.objects.create(
            name=payload["file"].name,
            file=payload["file"],
            description=payload.get("notes", ""),
            organization_id=payload.get("organization"),
            created_by=request.user,
        )
        job = ImportJob.objects.create(upload=upload, created_by=request.user, status="pending", started_at=timezone.now())
        _set_payload(
            job,
            SESSION_KEY,
            {
                "project": payload.get("project"),
                "organization": payload.get("organization"),
                "reporting_period": payload.get("reporting_period"),
                "template_id": payload.get("template_id"),
                "notes": payload.get("notes", ""),
                "status": "uploaded",
            },
        )
        _set_payload(job, ANALYSIS_KEY, {})
        job.save(update_fields=["errors", "started_at"])
        return Response(_serialize_session(job), status=status.HTTP_201_CREATED)

    @action(detail=True, methods=["post"])
    def analyze(self, request, pk=None):
        job = self.get_object()
        analysis = _analyze_workbook(job)
        template = _upsert_template_from_analysis(job, analysis)
        session = _get_payload(job, SESSION_KEY)
        session["template_id"] = template.id
        _set_payload(job, SESSION_KEY, session)
        _set_payload(job, ANALYSIS_KEY, analysis)
        _update_status(job, "ready_for_review" if not analysis["summary"]["errors"] else "failed")
        job.status = "completed" if not analysis["summary"]["errors"] else "failed"
        job.total_rows = analysis["summary"]["rows_imported"] + analysis["summary"]["rows_skipped"]
        job.processed_rows = job.total_rows
        job.successful_rows = analysis["summary"]["rows_imported"]
        job.failed_rows = analysis["summary"]["rows_skipped"]
        job.completed_at = timezone.now()
        job.save(update_fields=["errors", "status", "total_rows", "processed_rows", "successful_rows", "failed_rows", "completed_at"])
        return Response(_serialize_session(job))

    @action(detail=True, methods=["post"], url_path="create-missing-indicators")
    def create_missing_indicators(self, request, pk=None):
        job = self.get_object()
        serializer = CreateMissingIndicatorsSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        _apply_missing_indicators(job, serializer.validated_data, request.user)
        analysis = _get_payload(job, ANALYSIS_KEY)
        created_keys = {item["temp_key"] for item in serializer.validated_data["indicators"]}
        analysis["missing_indicators"] = [item for item in analysis.get("missing_indicators", []) if item.get("temp_key") not in created_keys]
        _set_payload(job, ANALYSIS_KEY, analysis)
        _update_status(job, "validated")
        job.save(update_fields=["errors"])
        return Response(_serialize_session(job))

    @action(detail=True, methods=["post"])
    def confirm(self, request, pk=None):
        job = self.get_object()
        serializer = ReportWorkbookConfirmSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        result = _confirm_import(job, serializer.validated_data, request.user)
        analysis = _get_payload(job, ANALYSIS_KEY)
        analysis["summary"] = result["summary"]
        analysis["issues"] = result["issues"]
        _set_payload(job, ANALYSIS_KEY, analysis)
        _update_status(job, "imported")
        job.status = "completed"
        job.total_rows = result["summary"]["rows_imported"] + result["summary"]["rows_skipped"]
        job.processed_rows = job.total_rows
        job.successful_rows = result["summary"]["rows_imported"]
        job.failed_rows = result["summary"]["rows_skipped"]
        job.completed_at = timezone.now()
        job.save(update_fields=["errors", "status", "total_rows", "processed_rows", "successful_rows", "failed_rows", "completed_at"])
        return Response(_serialize_session(job))


class ReportWorkbookTemplateViewSet(viewsets.ViewSet):
    permission_classes = [IsAuthenticated]

    def list(self, request):
        queryset = WorkbookTemplate.objects.filter(is_active=True).order_by("name", "-version")
        search = str(request.query_params.get("search") or "").strip().lower()
        workbook_family = str(request.query_params.get("workbook_family") or "").strip().lower()
        report_category = str(request.query_params.get("report_category") or "").strip().lower()
        if search:
            queryset = queryset.filter(name__icontains=search)
        if workbook_family:
            queryset = queryset.filter(workbook_family__iexact=workbook_family)
        if report_category:
            queryset = queryset.filter(report_category__iexact=report_category)
        data = [_serialize_template(item) for item in queryset]
        return Response({"count": len(data), "next": None, "previous": None, "results": data})

    def retrieve(self, request, pk=None):
        template = WorkbookTemplate.objects.filter(pk=pk).first()
        if not template:
            raise Http404("Workbook template not found.")
        return Response(_serialize_template(template))


class ReportWorkbookExportViewSet(viewsets.ViewSet):
    permission_classes = [IsAuthenticated]

    def create(self, request):
        serializer = ReportWorkbookExportCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        payload = serializer.validated_data

        project = Project.objects.filter(id=payload["project"]).first()
        if not project:
            return Response({"detail": "Project not found."}, status=status.HTTP_404_NOT_FOUND)

        template = WorkbookTemplate.objects.filter(id=payload.get("template_id")).first() if payload.get("template_id") else None
        organization = Organization.objects.filter(id=(payload.get("organization_ids") or [None])[0]).first() if payload.get("organization_ids") else None
        coordinator = Organization.objects.filter(id=payload.get("coordinator_id")).first() if payload.get("coordinator_id") else None

        job = WorkbookExportJob.objects.create(
            status="processing",
            scope=payload["scope"],
            reporting_period=payload["reporting_period"],
            financial_year_start_month=payload.get("financial_year_start_month", 4),
            project=project,
            template=template,
            organization=organization,
            coordinator=coordinator,
            created_by=request.user,
        )
        try:
            generated_upload = _build_export_workbook(
                job,
                include_validation_summary=payload.get("include_validation_summary", True),
            )
            job.generated_upload = generated_upload
            job.status = "completed"
            job.completed_at = timezone.now()
            job.save(update_fields=["generated_upload", "status", "completed_at"])
        except Exception as exc:  # pragma: no cover
            job.status = "failed"
            job.errors = [{"severity": "error", "code": "export_failed", "message": str(exc)}]
            job.completed_at = timezone.now()
            job.save(update_fields=["status", "errors", "completed_at"])
        return Response(_serialize_export_job(job, request=request), status=status.HTTP_201_CREATED)

    def retrieve(self, request, pk=None):
        job = WorkbookExportJob.objects.filter(pk=pk).select_related("generated_upload", "project").first()
        if not job:
            raise Http404("Workbook export job not found.")
        return Response(_serialize_export_job(job, request=request))

    @action(detail=True, methods=["get"])
    def download(self, request, pk=None):
        job = WorkbookExportJob.objects.filter(pk=pk).select_related("generated_upload").first()
        if not job or not job.generated_upload or not job.generated_upload.file:
            raise Http404("Workbook export file not found.")
        response = HttpResponse(
            job.generated_upload.file.open("rb"),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{Path(job.generated_upload.file.name).name}"'
        return response
