import json
import os
import re
import subprocess
import sys
from pathlib import Path

from django.conf import settings
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import MultiPartParser, FormParser
from django_filters.rest_framework import DjangoFilterBackend
from django.utils import timezone
from rest_framework import viewsets, status
from rest_framework.response import Response
from openpyxl import load_workbook

from .models import Upload, ImportJob
from .serializers import UploadSerializer, ImportJobSerializer


REPORT_WORKBOOK_IMPORT_SCRIPT = "import_selected_q3_workbook.py"
NON_REPORT_SHEET_KEYWORDS = (
    "indicator matrix",
    "summary",
    "instruction",
    "instructions",
    "cover",
    "contents",
    "readme",
    "notes",
)


def _is_truthy(value):
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {"1", "true", "yes", "on"}


def _parse_reporting_period_range(label):
    match = re.match(
        r"^Q([1-4])\s+(\d{4})(?:\s*/\s*(\d{2}|\d{4}))?$",
        str(label or "").strip(),
        flags=re.IGNORECASE,
    )
    if not match:
        return None

    quarter = int(match.group(1))
    year = int(match.group(2))
    next_year = year + 1

    if quarter == 1:
        return f"{year}-04-01", f"{year}-06-30"
    if quarter == 2:
        return f"{year}-07-01", f"{year}-09-30"
    if quarter == 3:
        return f"{year}-10-01", f"{year}-12-31"
    if quarter == 4:
        return f"{next_year}-01-01", f"{next_year}-03-31"
    return None


def _looks_like_non_reporting_sheet(sheet_name):
    normalized = " ".join(str(sheet_name or "").replace("_", " ").split()).lower()
    return any(keyword in normalized for keyword in NON_REPORT_SHEET_KEYWORDS)


def _resolve_report_workbook_import_script():
    candidates = [
        Path(settings.BASE_DIR) / "workbook-imports" / REPORT_WORKBOOK_IMPORT_SCRIPT,
        Path(settings.BASE_DIR).parent / "frontend" / "scripts" / REPORT_WORKBOOK_IMPORT_SCRIPT,
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return candidates[0]


def _get_importable_sheet_names(upload, provided_sheet_names=None):
    sheet_names = []

    if isinstance(provided_sheet_names, (list, tuple)):
        for sheet_name in provided_sheet_names:
            text = str(sheet_name or "").strip()
            if text and not _looks_like_non_reporting_sheet(text) and text not in sheet_names:
                sheet_names.append(text)
        if sheet_names:
            return sheet_names

    workbook = load_workbook(upload.file.path, read_only=True, data_only=True)
    try:
        for sheet_name in workbook.sheetnames:
            if _looks_like_non_reporting_sheet(sheet_name):
                continue
            if sheet_name not in sheet_names:
                sheet_names.append(sheet_name)
    finally:
        workbook.close()

    return sheet_names


def _coerce_mapping_payload(raw_value):
    if isinstance(raw_value, dict):
        return raw_value
    if isinstance(raw_value, str):
        text = raw_value.strip()
        if not text:
            return {}
        try:
            parsed = json.loads(text)
            if isinstance(parsed, dict):
                return parsed
        except Exception:
            return {}
    return {}


class UploadViewSet(viewsets.ModelViewSet):
    """ViewSet for managing uploads."""
    
    queryset = Upload.objects.all()
    serializer_class = UploadSerializer
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['file_type', 'organization', 'content_type']
    
    def get_queryset(self):
        user = self.request.user
        if user.is_superuser or user.is_staff or user.role == 'admin':
            return Upload.objects.all()
        elif user.organization:
            return Upload.objects.filter(organization=user.organization)
        return Upload.objects.filter(created_by=user)
    
    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)
    
    @action(detail=True, methods=['post'])
    def start_import(self, request, pk=None):
        """Start an import job for the uploaded file."""
        upload = self.get_object()
        queue_aggregate_review = _is_truthy(request.data.get("queue_aggregate_review"))
        dry_run = _is_truthy(request.data.get("dry_run"))

        job_status = "processing" if queue_aggregate_review else "ready_for_review"
        job = ImportJob.objects.create(
            upload=upload,
            status=job_status,
            started_at=timezone.now(),
            created_by=request.user
        )

        if not queue_aggregate_review:
            return Response(ImportJobSerializer(job).data, status=status.HTTP_201_CREATED)

        project_id = request.data.get("project_id") or request.data.get("project")
        try:
            project_id = int(project_id)
        except (TypeError, ValueError):
            project_id = None
        reporting_period = request.data.get("reporting_period") or request.data.get("period_label")
        period_start = request.data.get("period_start")
        period_end = request.data.get("period_end")
        if reporting_period and (not period_start or not period_end):
            parsed_period = _parse_reporting_period_range(reporting_period)
            if parsed_period:
                period_start, period_end = parsed_period

        if not project_id:
            job.status = "failed"
            job.completed_at = timezone.now()
            job.errors = [{"error": "project_id required for aggregate review queueing"}]
            job.save(update_fields=["status", "completed_at", "errors"])
            return Response(
                {"error": "project_id required for aggregate review queueing"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not period_start or not period_end:
            job.status = "failed"
            job.completed_at = timezone.now()
            job.errors = [{"error": "reporting_period or period_start/period_end required"}]
            job.save(update_fields=["status", "completed_at", "errors"])
            return Response(
                {"error": "reporting_period or period_start/period_end required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            sheet_names = _get_importable_sheet_names(upload, request.data.get("sheet_names"))
        except Exception as exc:
            job.status = "failed"
            job.completed_at = timezone.now()
            job.errors = [{"error": str(exc)}]
            job.save(update_fields=["status", "completed_at", "errors"])
            return Response(
                {"error": f"Unable to inspect workbook sheets: {exc}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not sheet_names:
            job.status = "failed"
            job.completed_at = timezone.now()
            job.errors = [{"error": "no importable organization sheets found"}]
            job.save(update_fields=["status", "completed_at", "errors"])
            return Response(
                {"error": "No importable organization sheets found in workbook"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        script_path = _resolve_report_workbook_import_script()
        if not script_path.exists():
            job.status = "failed"
            job.completed_at = timezone.now()
            job.errors = [{"error": f"Import script not found: {script_path}"}]
            job.save(update_fields=["status", "completed_at", "errors"])
            return Response(
                {"error": f"Import script not found: {script_path}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        report_path = Path(settings.BASE_DIR) / "reports" / "aggregate-review-queue" / f"upload-{upload.id}-job-{job.id}.json"
        command = [
            sys.executable,
            str(script_path),
            "--workbook",
            upload.file.path,
            "--project-id",
            str(project_id),
            "--period-start",
            str(period_start),
            "--period-end",
            str(period_end),
            "--report-path",
            str(report_path),
            "--sheets",
            *sheet_names,
        ]
        if dry_run:
            command.append("--dry-run")

        indicator_overrides = _coerce_mapping_payload(request.data.get("indicator_overrides"))
        if indicator_overrides:
            indicator_overrides_path = report_path.with_suffix(".indicator-overrides.json")
            indicator_overrides_path.parent.mkdir(parents=True, exist_ok=True)
            indicator_overrides_path.write_text(
                json.dumps(indicator_overrides, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
            command.extend(["--indicator-overrides-path", str(indicator_overrides_path)])

        sheet_org_overrides = _coerce_mapping_payload(request.data.get("sheet_org_overrides"))
        if sheet_org_overrides:
            sheet_org_overrides_path = report_path.with_suffix(".sheet-org-overrides.json")
            sheet_org_overrides_path.parent.mkdir(parents=True, exist_ok=True)
            sheet_org_overrides_path.write_text(
                json.dumps(sheet_org_overrides, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
            command.extend(["--sheet-org-overrides-path", str(sheet_org_overrides_path)])

        env = os.environ.copy()
        env["BONASO_DJANGO_ROOT"] = str(settings.BASE_DIR)

        try:
            subprocess.run(
                command,
                cwd=str(settings.BASE_DIR),
                env=env,
                check=True,
                capture_output=True,
                text=True,
            )
        except subprocess.CalledProcessError as exc:
            error_message = (exc.stderr or exc.stdout or str(exc)).strip()
            unresolved_sheet_names = []
            unresolved_match = re.search(
                r"Organizations not found for sheets:\s*(.+)$",
                error_message,
                flags=re.IGNORECASE | re.MULTILINE,
            )
            if unresolved_match:
                unresolved_sheet_names = [
                    item.strip()
                    for item in unresolved_match.group(1).split(",")
                    if item.strip()
                ]
            job.status = "failed"
            job.completed_at = timezone.now()
            job.errors = [
                {
                    "error": "aggregate import failed",
                    "details": error_message,
                    "unresolved_sheet_names": unresolved_sheet_names,
                }
            ]
            job.save(update_fields=["status", "completed_at", "errors"])
            return Response(
                {
                    "error": "Aggregate import failed",
                    "details": error_message,
                    "unresolved_sheet_names": unresolved_sheet_names,
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        summary = {}
        report_payload = {}
        try:
            if report_path.exists():
                report_payload = json.loads(report_path.read_text(encoding="utf-8"))
                summary = report_payload.get("summary", {})
        except Exception:
            summary = {}
            report_payload = {}

        matched_rows = int(summary.get("matched_rows", 0) or 0)
        unknown_rows = int(summary.get("unknown_rows", 0) or 0)
        total_rows = matched_rows + unknown_rows

        job.status = "validated" if dry_run else "imported"
        job.total_rows = total_rows
        job.processed_rows = matched_rows
        job.successful_rows = matched_rows
        job.failed_rows = unknown_rows
        job.completed_at = timezone.now()
        job.errors = []
        job.save(
            update_fields=[
                "status",
                "total_rows",
                "processed_rows",
                "successful_rows",
                "failed_rows",
                "completed_at",
                "errors",
            ]
        )

        payload = ImportJobSerializer(job).data
        payload["aggregate_review_queued"] = not dry_run
        payload["dry_run"] = dry_run
        payload["aggregate_import_summary"] = summary
        payload["aggregate_import_report"] = report_payload
        return Response(payload, status=status.HTTP_201_CREATED)


class ImportJobViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet for viewing import jobs."""
    
    queryset = ImportJob.objects.all()
    serializer_class = ImportJobSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['status', 'upload']
    
    def get_queryset(self):
        user = self.request.user
        if user.is_superuser or user.is_staff or user.role == 'admin':
            return ImportJob.objects.all()
        return ImportJob.objects.filter(created_by=user)

