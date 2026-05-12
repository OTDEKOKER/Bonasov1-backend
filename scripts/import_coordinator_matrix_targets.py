import argparse
import json
import os
import re
import sys
import zipfile
from decimal import Decimal
from io import BytesIO
from pathlib import Path


BACKEND_ROOT = Path(__file__).resolve().parents[1]
if str(BACKEND_ROOT) not in sys.path:
    sys.path.insert(0, str(BACKEND_ROOT))

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "core.settings")

import django  # noqa: E402

django.setup()

from django.db import connection, transaction  # noqa: E402

from organizations.models import Organization  # noqa: E402
from projects.models import Project  # noqa: E402
from uploads.management.commands.import_reporting_workbook_overwrite import (  # noqa: E402
    IndicatorResolver,
    OrganizationResolver,
    ensure_project_indicator,
    find_matrix_sheet_name,
    parse_matrix_sheet,
)


WORKBOOK_EXTENSIONS = {".xlsx", ".xlsm", ".xls"}
IGNORED_NAME_TOKENS = ("matrix-total-values-import", "total-sheet-import")

COORDINATOR_DEFINITIONS = [
    {
        "key": "BONELA",
        "folder_hints": ["BONELA"],
        "organization_aliases": ["BONELA"],
    },
    {
        "key": "BONEPWA",
        "folder_hints": ["BONEPWA+"],
        "organization_aliases": ["BONEPWA", "BONEPWA+"],
    },
    {
        "key": "MAKGABANENG",
        "folder_hints": ["MAKGABANENG"],
        "organization_aliases": ["MAKGABANENG"],
    },
    {
        "key": "MBGE",
        "folder_hints": ["MBGE"],
        "organization_aliases": ["MBGE", "Men for Health and Gender Justice Org."],
    },
    {
        "key": "TEBELOPELE",
        "folder_hints": ["TEBELOPELE"],
        "organization_aliases": ["TEBELOPELE"],
    },
]


def to_decimal(value) -> Decimal:
    if isinstance(value, Decimal):
        return value
    if value in (None, ""):
        return Decimal("0")
    if isinstance(value, (int, float)):
        return Decimal(str(value))

    text = str(value).strip().replace(",", "")
    if not text:
        return Decimal("0")
    try:
        return Decimal(text)
    except Exception:
        match = re.search(r"-?\d+(?:\.\d+)?", text)
        if not match:
            return Decimal("0")
        try:
            return Decimal(match.group(0))
        except Exception:
            return Decimal("0")


def normalize_token(value) -> str:
    return "".join(ch.lower() if ch.isalnum() else " " for ch in str(value or "")).strip()


def workbook_name_score(name: str) -> tuple[int, str]:
    lowered = name.lower()
    score = 0
    if lowered.startswith("~$"):
        score += 500
    if "normalized" in lowered:
        score += 25
    if any(token in lowered for token in IGNORED_NAME_TOKENS):
        score += 150
    if Path(name).suffix.lower() not in WORKBOOK_EXTENSIONS:
        score += 1000
    return score, lowered


def select_workbook_from_folder(folder: Path) -> dict | None:
    if not folder.exists() or not folder.is_dir():
        return None

    direct_candidates = []
    zip_candidates = []

    for child in folder.iterdir():
        if child.is_file() and child.suffix.lower() in WORKBOOK_EXTENSIONS:
            direct_candidates.append((workbook_name_score(child.name), child))
            continue
        if child.is_file() and child.suffix.lower() == ".zip":
            try:
                with zipfile.ZipFile(child) as archive:
                    for member in archive.namelist():
                        if member.endswith("/"):
                            continue
                        member_name = Path(member).name
                        if Path(member_name).suffix.lower() not in WORKBOOK_EXTENSIONS:
                            continue
                        zip_candidates.append((workbook_name_score(member_name), child, member))
            except zipfile.BadZipFile:
                continue

    if direct_candidates:
        direct_candidates.sort(key=lambda item: item[0])
        _, path = direct_candidates[0]
        return {
            "type": "file",
            "display_name": path.name,
            "path": str(path),
        }

    if zip_candidates:
        zip_candidates.sort(key=lambda item: item[0])
        _, zip_path, member = zip_candidates[0]
        return {
            "type": "zip-entry",
            "display_name": Path(member).name,
            "path": str(zip_path),
            "zip_member": member,
        }

    return None


def load_workbook_from_selection(selection: dict):
    from openpyxl import load_workbook

    if selection["type"] == "file":
        return load_workbook(Path(selection["path"]), data_only=True, read_only=True, keep_links=False)

    zip_path = Path(selection["path"])
    with zipfile.ZipFile(zip_path) as archive:
        payload = archive.read(selection["zip_member"])
    return load_workbook(BytesIO(payload), data_only=True, read_only=True, keep_links=False)


def parse_matrix_sheet_fallback(ws, coordinator_name: str) -> dict[str, dict]:
    rows = list(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 24), min_col=1, max_col=min(ws.max_column, 40), values_only=True))
    header_row_index = None
    indicator_col = None
    q1_col = None
    q2_col = None
    q3_col = None
    q4_col = None
    org_col = None

    for row_idx, row_values in enumerate(rows, start=1):
        token_by_col = {}
        for col_idx, value in enumerate(row_values, start=1):
            token = normalize_token(value)
            if token:
                token_by_col[col_idx] = token

        indicator_col_candidate = next((c for c, t in token_by_col.items() if t == "indicator" or "indicator" in t), None)
        q1_candidate = next((c for c, t in token_by_col.items() if t.startswith("q1")), None)
        q2_candidate = next((c for c, t in token_by_col.items() if t.startswith("q2")), None)
        q3_candidate = next((c for c, t in token_by_col.items() if t.startswith("q3")), None)
        q4_candidate = next((c for c, t in token_by_col.items() if t.startswith("q4")), None)
        org_candidate = next((c for c, t in token_by_col.items() if "organization" in t or t == "org"), None)

        if indicator_col_candidate and q1_candidate and q2_candidate and q3_candidate and q4_candidate:
            header_row_index = row_idx
            indicator_col = indicator_col_candidate
            q1_col = q1_candidate
            q2_col = q2_candidate
            q3_col = q3_candidate
            q4_col = q4_candidate
            org_col = org_candidate
            break

    if header_row_index is None:
        return {}

    assignments = {}
    current_org_name = coordinator_name

    for row_values in ws.iter_rows(
        min_row=header_row_index + 1,
        min_col=1,
        max_col=min(ws.max_column, 40),
        values_only=True,
    ):
        indicator_value = row_values[indicator_col - 1] if indicator_col - 1 < len(row_values) else None
        indicator_title = str(indicator_value or "").strip()
        if not indicator_title:
            continue
        indicator_token = normalize_token(indicator_title)
        if indicator_token in {"indicator", "numerator", "denominator"}:
            continue

        if org_col:
            org_value = row_values[org_col - 1] if org_col - 1 < len(row_values) else None
            if str(org_value or "").strip():
                current_org_name = str(org_value).strip()

        q1 = to_decimal(row_values[q1_col - 1] if q1_col - 1 < len(row_values) else 0)
        q2 = to_decimal(row_values[q2_col - 1] if q2_col - 1 < len(row_values) else 0)
        q3 = to_decimal(row_values[q3_col - 1] if q3_col - 1 < len(row_values) else 0)
        q4 = to_decimal(row_values[q4_col - 1] if q4_col - 1 < len(row_values) else 0)
        if q1 == 0 and q2 == 0 and q3 == 0 and q4 == 0:
            continue

        key = indicator_title.lower().strip()
        entry = assignments.setdefault(
            key,
            {
                "title": indicator_title,
                "assignments": [],
            },
        )
        entry["assignments"].append(
            {
                "organization_name": current_org_name,
                "q1_target": q1,
                "q2_target": q2,
                "q3_target": q3,
                "q4_target": q4,
            }
        )

    return assignments


def count_titles_with_letters(assignments: dict[str, dict]) -> int:
    count = 0
    for assignment_bundle in assignments.values():
        title = str(assignment_bundle.get("title", ""))
        if any(ch.isalpha() for ch in title):
            count += 1
    return count


def resolve_project(project_id: int, project_code: str) -> Project:
    project = None
    if project_id:
        project = Project.objects.filter(id=project_id).first()
    if project is None and project_code:
        project = Project.objects.filter(code__iexact=project_code).first()
    if project is None:
        raise SystemExit("Project not found. Provide --project-id or --project-code.")
    return project


def resolve_coordinator(aliases: list[str]) -> Organization | None:
    for alias in aliases:
        by_code = Organization.objects.filter(code__iexact=alias).first()
        if by_code:
            return by_code
        by_name = Organization.objects.filter(name__iexact=alias).first()
        if by_name:
            return by_name
    return None


def belongs_to_coordinator(
    organization_id: int,
    coordinator_id: int,
    parent_by_org_id: dict[int, int | None],
) -> bool:
    if organization_id == coordinator_id:
        return True

    seen = set()
    current = organization_id
    while current and current not in seen:
        seen.add(current)
        parent_id = parent_by_org_id.get(current)
        if parent_id == coordinator_id:
            return True
        current = parent_id
    return False


def upsert_org_target_row(
    cursor,
    project_indicator_id: int,
    organization_id: int,
    q1_target: Decimal,
    q2_target: Decimal,
    q3_target: Decimal,
    q4_target: Decimal,
    dry_run: bool,
) -> str:
    target_value = q1_target + q2_target + q3_target + q4_target
    cursor.execute(
        """
        SELECT id, q1_target, q2_target, q3_target, q4_target, target_value
        FROM projects_projectindicatororganizationtarget
        WHERE project_indicator_id = %s AND organization_id = %s
        """,
        [project_indicator_id, organization_id],
    )
    row = cursor.fetchone()

    if row is None:
        if not dry_run:
            cursor.execute(
                """
                INSERT INTO projects_projectindicatororganizationtarget
                    (q1_target, q2_target, q3_target, q4_target, target_value, current_value, baseline_value, organization_id, project_indicator_id)
                VALUES (%s, %s, %s, %s, %s, 0, 0, %s, %s)
                """,
                [
                    q1_target,
                    q2_target,
                    q3_target,
                    q4_target,
                    target_value,
                    organization_id,
                    project_indicator_id,
                ],
            )
        return "created"

    current_values = tuple(to_decimal(value) for value in row[1:6])
    desired_values = (q1_target, q2_target, q3_target, q4_target, target_value)
    if current_values == desired_values:
        return "unchanged"

    if not dry_run:
        cursor.execute(
            """
            UPDATE projects_projectindicatororganizationtarget
            SET q1_target = %s,
                q2_target = %s,
                q3_target = %s,
                q4_target = %s,
                target_value = %s
            WHERE id = %s
            """,
            [q1_target, q2_target, q3_target, q4_target, target_value, row[0]],
        )
    return "updated"


def sync_project_indicator_totals(
    cursor,
    project_indicator_id: int,
    dry_run: bool,
) -> str:
    cursor.execute(
        """
        SELECT
            COALESCE(SUM(q1_target), 0),
            COALESCE(SUM(q2_target), 0),
            COALESCE(SUM(q3_target), 0),
            COALESCE(SUM(q4_target), 0),
            COALESCE(SUM(target_value), 0)
        FROM projects_projectindicatororganizationtarget
        WHERE project_indicator_id = %s
        """,
        [project_indicator_id],
    )
    summed = tuple(to_decimal(value) for value in cursor.fetchone())

    cursor.execute(
        """
        SELECT q1_target, q2_target, q3_target, q4_target, target_value
        FROM projects_projectindicator
        WHERE id = %s
        """,
        [project_indicator_id],
    )
    row = cursor.fetchone()
    if row is None:
        return "missing_project_indicator"

    current = tuple(to_decimal(value) for value in row)
    if current == summed:
        return "unchanged"

    if not dry_run:
        cursor.execute(
            """
            UPDATE projects_projectindicator
            SET q1_target = %s,
                q2_target = %s,
                q3_target = %s,
                q4_target = %s,
                target_value = %s
            WHERE id = %s
            """,
            [summed[0], summed[1], summed[2], summed[3], summed[4], project_indicator_id],
        )
    return "updated"


def build_args():
    parser = argparse.ArgumentParser(
        description=(
            "Choose one workbook per coordinator folder and import "
            "Indicator Matrix quarterly targets (Q1-Q4) into organization targets."
        )
    )
    parser.add_argument(
        "--source-root",
        default="/home/bonasoadmin/BONASOV1/imports/Q4 CSO REPORTS/Q4 CSO REPORTS",
    )
    parser.add_argument("--project-id", type=int, default=0)
    parser.add_argument("--project-code", default="NAHPA2025/26")
    parser.add_argument("--coordinators", nargs="*", default=[])
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--report-path", default="")
    return parser.parse_args()


def main():
    args = build_args()
    source_root = Path(args.source_root).expanduser()
    if not source_root.exists():
        raise SystemExit(f"Source root not found: {source_root}")

    project = resolve_project(args.project_id, args.project_code)
    org_resolver = OrganizationResolver()
    indicator_resolver = IndicatorResolver(project=project)

    requested = {key.upper() for key in args.coordinators if key.strip()}
    coordinator_defs = [
        item
        for item in COORDINATOR_DEFINITIONS
        if not requested or item["key"] in requested
    ]
    if not coordinator_defs:
        raise SystemExit("No coordinator definitions matched --coordinators.")

    parent_by_org_id = dict(
        Organization.objects.values_list("id", "parent_id")
    )

    report = {
        "source_root": str(source_root),
        "project": {"id": project.id, "code": project.code, "name": project.name},
        "dry_run": args.dry_run,
        "selections": [],
        "coordinators": {},
        "summary": {
            "coordinators_requested": len(coordinator_defs),
            "coordinators_selected": 0,
            "matrix_rows_seen": 0,
            "assignments_seen": 0,
            "targets_created": 0,
            "targets_updated": 0,
            "targets_unchanged": 0,
            "project_totals_updated": 0,
            "project_totals_unchanged": 0,
            "missing_indicator_rows": 0,
            "missing_organizations": 0,
            "assignments_outside_coordinator_tree": 0,
        },
    }

    selections = []
    for definition in coordinator_defs:
        coordinator = resolve_coordinator(definition["organization_aliases"])
        if not coordinator:
            report["coordinators"][definition["key"]] = {
                "status": "coordinator_not_found",
                "organization_aliases": definition["organization_aliases"],
            }
            continue

        selection = None
        for folder_hint in definition["folder_hints"]:
            candidate = select_workbook_from_folder(source_root / folder_hint)
            if candidate:
                selection = candidate
                break

        if not selection:
            report["coordinators"][definition["key"]] = {
                "status": "workbook_not_found",
                "coordinator_id": coordinator.id,
                "coordinator_name": coordinator.name,
                "folder_hints": definition["folder_hints"],
            }
            continue

        selections.append((definition["key"], coordinator, selection))
        report["summary"]["coordinators_selected"] += 1
        report["selections"].append(
            {
                "coordinator_key": definition["key"],
                "coordinator_id": coordinator.id,
                "coordinator_name": coordinator.name,
                **selection,
            }
        )

    touched_project_indicator_ids = set()

    with transaction.atomic():
        with connection.cursor() as cursor:
            for coordinator_key, coordinator, selection in selections:
                coordinator_report = {
                    "status": "ok",
                    "coordinator_id": coordinator.id,
                    "coordinator_name": coordinator.name,
                    "workbook": selection,
                    "matrix_rows_seen": 0,
                    "assignments_seen": 0,
                    "targets_created": 0,
                    "targets_updated": 0,
                    "targets_unchanged": 0,
                    "missing_indicator_rows": [],
                    "missing_organizations": [],
                    "assignments_outside_coordinator_tree": [],
                }
                report["coordinators"][coordinator_key] = coordinator_report

                workbook = load_workbook_from_selection(selection)
                matrix_sheet_name = find_matrix_sheet_name(list(workbook.sheetnames))
                if not matrix_sheet_name:
                    coordinator_report["status"] = "missing_matrix_sheet"
                    continue

                matrix_sheet = workbook[matrix_sheet_name]
                matrix_assignments = parse_matrix_sheet(matrix_sheet)
                fallback_assignments = parse_matrix_sheet_fallback(matrix_sheet, coordinator.name)
                primary_title_score = count_titles_with_letters(matrix_assignments)
                fallback_title_score = count_titles_with_letters(fallback_assignments)
                if fallback_title_score > primary_title_score:
                    matrix_assignments = fallback_assignments
                    coordinator_report["parser_mode"] = "fallback_dynamic_headers"
                else:
                    coordinator_report["parser_mode"] = "default_matrix_parser"

                coordinator_report["matrix_rows_seen"] = len(matrix_assignments)
                report["summary"]["matrix_rows_seen"] += len(matrix_assignments)

                for assignment_bundle in matrix_assignments.values():
                    title = assignment_bundle.get("title", "")
                    indicator = indicator_resolver.resolve(title, None)
                    if not indicator:
                        coordinator_report["missing_indicator_rows"].append(title)
                        report["summary"]["missing_indicator_rows"] += 1
                        continue

                    project_indicator, _ = ensure_project_indicator(project=project, indicator=indicator)
                    touched_project_indicator_ids.add(project_indicator.id)

                    for assignment in assignment_bundle.get("assignments", []):
                        coordinator_report["assignments_seen"] += 1
                        report["summary"]["assignments_seen"] += 1

                        assigned_org = org_resolver.resolve(assignment.get("organization_name"))
                        if not assigned_org:
                            coordinator_report["missing_organizations"].append(
                                assignment.get("organization_name")
                            )
                            report["summary"]["missing_organizations"] += 1
                            continue

                        in_tree = belongs_to_coordinator(
                            organization_id=assigned_org.id,
                            coordinator_id=coordinator.id,
                            parent_by_org_id=parent_by_org_id,
                        )
                        if not in_tree:
                            coordinator_report["assignments_outside_coordinator_tree"].append(
                                {
                                    "organization_id": assigned_org.id,
                                    "organization_name": assigned_org.name,
                                    "matrix_organization_name": assignment.get("organization_name"),
                                }
                            )
                            report["summary"]["assignments_outside_coordinator_tree"] += 1

                        project.organizations.add(assigned_org)
                        indicator.organizations.add(assigned_org)

                        q1 = to_decimal(assignment.get("q1_target"))
                        q2 = to_decimal(assignment.get("q2_target"))
                        q3 = to_decimal(assignment.get("q3_target"))
                        q4 = to_decimal(assignment.get("q4_target"))
                        action = upsert_org_target_row(
                            cursor=cursor,
                            project_indicator_id=project_indicator.id,
                            organization_id=assigned_org.id,
                            q1_target=q1,
                            q2_target=q2,
                            q3_target=q3,
                            q4_target=q4,
                            dry_run=args.dry_run,
                        )
                        if action == "created":
                            coordinator_report["targets_created"] += 1
                            report["summary"]["targets_created"] += 1
                        elif action == "updated":
                            coordinator_report["targets_updated"] += 1
                            report["summary"]["targets_updated"] += 1
                        else:
                            coordinator_report["targets_unchanged"] += 1
                            report["summary"]["targets_unchanged"] += 1

            for project_indicator_id in touched_project_indicator_ids:
                action = sync_project_indicator_totals(
                    cursor=cursor,
                    project_indicator_id=project_indicator_id,
                    dry_run=args.dry_run,
                )
                if action == "updated":
                    report["summary"]["project_totals_updated"] += 1
                elif action == "unchanged":
                    report["summary"]["project_totals_unchanged"] += 1

        if args.dry_run:
            transaction.set_rollback(True)

    report_path = (
        Path(args.report_path)
        if args.report_path
        else BACKEND_ROOT / "reports" / "coordinator-matrix-target-import-report.json"
    )
    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text(json.dumps(report, indent=2), encoding="utf-8")

    print(json.dumps(report["summary"], indent=2))
    print(f"Saved report: {report_path}")


if __name__ == "__main__":
    main()
