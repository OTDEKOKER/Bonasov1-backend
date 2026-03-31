"""
Import coordinator quarterly targets from the Social Contracting workbook.

Usage:
  python scripts/import_coordinator_targets_from_workbook.py

This script is idempotent:
- existing coordinator-target rows are updated when values change
- new rows are created when missing
"""

from __future__ import annotations

import difflib
import os
import re
import sys
from collections import defaultdict
from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path

import openpyxl

ROOT_DIR = Path(__file__).resolve().parents[1]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "core.settings")

import django  # noqa: E402

django.setup()

from analysis.models import CoordinatorTarget  # noqa: E402
from indicators.models import Indicator  # noqa: E402
from organizations.models import Organization  # noqa: E402
from projects.models import Project  # noqa: E402


WORKBOOK_PATH = Path(
    r"C:\Users\dekok\Downloads\Social Contracting 2025-Final Targets & indicators 5 Clusters 14 AUGUST 202526.xlsx"
)
FISCAL_YEAR = 2025

# Explicit request: skip CLUSTER 6-BBCA.
SKIP_SHEETS = {"INSTRUCTIONS", "CLUSTER 6-BBCA"}

SHEET_TO_COORDINATOR = {
    "CLUSTER 1-BONELA": "BONELA",
    "CLUSTER 2-Tebelopele": "TEBELOPELE",
    "CLUSTER 3-BONEPWA+": "BONEPWA",
    "CLUSTER 4-Men & Boys": "MBGE",
    "CLUSTER 5-Makgabaneng": "MAKGABANENG",
    "Cluster 7-BONASO": "BONASO",
}

# Workbook typos/variant names -> canonical indicator names in DB (normalized).
INDICATOR_ALIAS_MAP = {
    "number of ayps linked to care": "number of ayp linked to care",
    "number of condoms distributed to ayp": "number of condoms distributed to ayps",
    "number of condoms distributed to plwh": "number of condoms distributed to kvps",
    "number of coodinators who attended orientation on project scope financing expectations and coodinatio roles": (
        "number of coodinators who attended orientation on project scope m e expectations and coordination roles"
    ),
    "number of people reached with hiv prevention messages": "reached with hiv prevention and control messages",
    "number of service providers receiving training community mobilisers m e officers program officers stakeholders": (
        "number of service providers receiving training community mobilisers m e officers stakeholders"
    ),
    "numberof lubricants distributed": "number of lubricants distributed",
    "proportion of functional tobacco cessation and alcohol abuse support group": (
        "number of functional tobacco cessation and alcohol abuse support groups"
    ),
}

QUARTERS = ("Q1", "Q2", "Q3", "Q4")
IGNORE_INDICATOR_MARKERS = {
    "indicator",
    "numerator",
    "denominator",
    "dissagregation",
    "disaggregation",
    "age",
    "m",
    "f",
    "coordinator",
}


@dataclass(frozen=True)
class TargetKey:
    coordinator_id: int
    indicator_id: int


def normalize_text(value: str | None) -> str:
    text = str(value or "").strip().lower()
    text = text.replace("’", "'").replace("“", '"').replace("”", '"')
    text = text.replace("&", " and ")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def to_decimal(value) -> Decimal:
    if value is None or value == "":
        return Decimal("0")
    if isinstance(value, bool):
        return Decimal("1") if value else Decimal("0")
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    raw = str(value).strip().replace(",", "")
    if not raw:
        return Decimal("0")
    try:
        return Decimal(raw)
    except Exception:
        return Decimal("0")


def resolve_project() -> Project:
    project = Project.objects.filter(name__icontains="social contracting").first()
    if project:
        return project
    project = Project.objects.filter(code__icontains="NAHPA").first()
    if project:
        return project
    project = Project.objects.first()
    if project:
        return project
    raise RuntimeError("No project found. Create a project before importing coordinator targets.")


def find_header(ws) -> tuple[int | None, int | None, dict[str, int]]:
    header_row = None
    indicator_col = None
    q_cols: dict[str, int] = {}

    scan_max = min(ws.max_row, 60)
    for row_index in range(1, scan_max + 1):
        row_values = [ws.cell(row_index, col).value for col in range(1, ws.max_column + 1)]
        row_text = [str(value).strip() if value is not None else "" for value in row_values]

        if not (any("Q1" in text for text in row_text) and any("Indicator" in text for text in row_text)):
            continue

        header_row = row_index
        for col_index, text in enumerate(row_text, start=1):
            lower_text = text.lower()
            if "q1" in lower_text:
                q_cols["Q1"] = col_index
            elif "q2" in lower_text:
                q_cols["Q2"] = col_index
            elif "q3" in lower_text:
                q_cols["Q3"] = col_index
            elif "q4" in lower_text:
                q_cols["Q4"] = col_index
            elif "indicator" in lower_text and indicator_col is None:
                indicator_col = col_index
        break

    return header_row, indicator_col, q_cols


def resolve_indicator(
    raw_name: str,
    indicator_by_norm: dict[str, Indicator],
) -> Indicator | None:
    normalized = normalize_text(raw_name)
    normalized = INDICATOR_ALIAS_MAP.get(normalized, normalized)

    direct = indicator_by_norm.get(normalized)
    if direct:
        return direct

    # Containment fallback.
    containment_matches = [indicator for key, indicator in indicator_by_norm.items() if normalized in key or key in normalized]
    if len(containment_matches) == 1:
        return containment_matches[0]

    # Similarity fallback (high confidence only).
    all_keys = list(indicator_by_norm.keys())
    close_matches = difflib.get_close_matches(normalized, all_keys, n=1, cutoff=0.9)
    if close_matches:
        return indicator_by_norm[close_matches[0]]

    return None


def main():
    if not WORKBOOK_PATH.exists():
        raise FileNotFoundError(f"Workbook not found: {WORKBOOK_PATH}")

    project = resolve_project()
    print(f"Project: {project.id} - {project.name}")
    print(f"Workbook: {WORKBOOK_PATH}")
    print(f"Fiscal year: {FISCAL_YEAR}")

    wb = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)

    organizations = list(Organization.objects.all())
    org_by_norm = {normalize_text(org.name): org for org in organizations}

    indicators = list(Indicator.objects.all())
    indicator_by_norm = {normalize_text(indicator.name): indicator for indicator in indicators}

    aggregated_targets: dict[TargetKey, dict[str, Decimal]] = defaultdict(
        lambda: {quarter: Decimal("0") for quarter in QUARTERS}
    )

    unmatched_sheets: list[str] = []
    unmatched_organizations: list[tuple[str, str]] = []
    unmatched_indicators: list[tuple[str, str]] = []

    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            print(f"Skip sheet: {sheet_name}")
            continue
        if sheet_name not in SHEET_TO_COORDINATOR:
            unmatched_sheets.append(sheet_name)
            continue

        coordinator_name = SHEET_TO_COORDINATOR[sheet_name]
        coordinator = org_by_norm.get(normalize_text(coordinator_name))
        if coordinator is None:
            unmatched_organizations.append((sheet_name, coordinator_name))
            continue

        ws = wb[sheet_name]
        header_row, indicator_col, q_cols = find_header(ws)
        if header_row is None or indicator_col is None or any(quarter not in q_cols for quarter in QUARTERS):
            unmatched_sheets.append(sheet_name)
            continue

        parsed_rows = 0
        matched_rows = 0
        for row_index in range(header_row + 1, ws.max_row + 1):
            raw_indicator_name = ws.cell(row_index, indicator_col).value
            if not isinstance(raw_indicator_name, str):
                continue
            raw_indicator_name = raw_indicator_name.strip()
            if not raw_indicator_name:
                continue
            if normalize_text(raw_indicator_name) in IGNORE_INDICATOR_MARKERS:
                continue

            quarter_values = {quarter: to_decimal(ws.cell(row_index, q_cols[quarter]).value) for quarter in QUARTERS}

            # Accept rows where any quarter cell is numeric-like (including zero).
            if not any(
                isinstance(ws.cell(row_index, q_cols[quarter]).value, (int, float, bool))
                or str(ws.cell(row_index, q_cols[quarter]).value or "").strip().replace(",", "").replace(".", "").isdigit()
                for quarter in QUARTERS
            ):
                continue

            parsed_rows += 1
            indicator = resolve_indicator(raw_indicator_name, indicator_by_norm)
            if indicator is None:
                unmatched_indicators.append((sheet_name, raw_indicator_name))
                continue

            matched_rows += 1
            key = TargetKey(coordinator_id=coordinator.id, indicator_id=indicator.id)
            for quarter in QUARTERS:
                aggregated_targets[key][quarter] += quarter_values[quarter]

        print(f"{sheet_name}: parsed {parsed_rows} indicator rows, matched {matched_rows}")

    created = 0
    updated = 0
    unchanged = 0

    for key, quarter_targets in aggregated_targets.items():
        for quarter in QUARTERS:
            target_value = quarter_targets[quarter]
            source_note = f"Imported from {WORKBOOK_PATH.name} ({FISCAL_YEAR})"

            target, was_created = CoordinatorTarget.objects.get_or_create(
                project=project,
                coordinator_id=key.coordinator_id,
                indicator_id=key.indicator_id,
                year=FISCAL_YEAR,
                quarter=quarter,
                defaults={
                    "target_value": target_value,
                    "notes": source_note,
                    "is_active": True,
                },
            )
            if was_created:
                created += 1
                continue

            changed = False
            if target.target_value != target_value:
                target.target_value = target_value
                changed = True
            if (target.notes or "") != source_note:
                target.notes = source_note
                changed = True
            if target.is_active is False:
                target.is_active = True
                changed = True

            if changed:
                target.save(update_fields=["target_value", "notes", "is_active", "updated_at"])
                updated += 1
            else:
                unchanged += 1

    unique_unmatched_indicators = sorted({entry for entry in unmatched_indicators})

    print("\nImport summary")
    print(f"- aggregated coordinator/indicator pairs: {len(aggregated_targets)}")
    print(f"- created quarter records: {created}")
    print(f"- updated quarter records: {updated}")
    print(f"- unchanged quarter records: {unchanged}")
    print(f"- unmatched sheets: {len(unmatched_sheets)}")
    for sheet in unmatched_sheets:
        print(f"  * {sheet}")
    print(f"- unmatched organizations: {len(unmatched_organizations)}")
    for sheet, org_name in unmatched_organizations:
        print(f"  * {sheet} -> {org_name}")
    print(f"- unmatched indicators: {len(unique_unmatched_indicators)}")
    for sheet, indicator_name in unique_unmatched_indicators:
        print(f"  * {sheet}: {indicator_name}")


if __name__ == "__main__":
    main()
